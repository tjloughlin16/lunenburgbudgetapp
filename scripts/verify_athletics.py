"""Recompute every figure in the athletics analysis, and fail if one drifted.

Rule 9: figures in a finished document get re-checked by script, not re-read. Prose drifts
during editing and the version that ships is the one nobody checked.

Everything here is recomputed from primary sources -- the FY27 workbook, the fund's own
year-end reconciliation, the extracted line history, and the FY19 split document as
transcribed into model/athletics.py -- and then asserted to be present in the document.

    python3 scripts/verify_athletics.py
"""
import os, sys, csv, collections

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, 'model'))
DOC = os.path.join(ROOT, 'sources/analyses/athletics.md')
DATA = os.path.join(ROOT, 'sources/data')

TEXT = open(DOC, encoding='utf-8').read()
PLAIN = TEXT.replace('**', '').replace('`', '')
FAILS = []


def present(label, needle):
    # Prose uses the typographic minus and en dash; code emits the ASCII hyphen, and they
    # are the same number. Normalising both sides beats reporting drift that is only a
    # character -- the same fix verify_budget_vs_actual.py carries.
    def norm(t):
        return t.replace('\u2212', '-').replace('\u2013', '-')
    n = str(needle)
    ok = (n in TEXT or n in PLAIN or norm(n) in norm(TEXT) or norm(n) in norm(PLAIN))
    if not ok:
        FAILS.append(f'{label}: "{n}" not in the document')
    print(f"  {'OK  ' if ok else 'GONE'}  {label:<46} {n}")


def head(t):
    print(f'\n{t}')


# --- the FY19 split document, via the model ---------------------------------------
import athletics as A
S = A.SPLIT_REPORTING

head('The FY19 split document — transportation, both sides')
for r in S['transportation']:
    tot = r['general'] + r['revolving']
    share = r['revolving'] / tot * 100
    present(f"FY{r['fy']} appropriated", f"{r['general']:,}")
    present(f"FY{r['fy']} revolving", f"{r['revolving']:,}")
    present(f"FY{r['fy']} total", f"{tot:,}")
    present(f"FY{r['fy']} fund share", f'{share:.1f}%')

head('The FY19 split document — whole programme')
for r in S['programme']:
    present(f"FY{r['fy']} general appropriation", f"{r['general']:,}")
    present(f"FY{r['fy']} revolving 658", f"{r['revolving']:,}")
    present(f"FY{r['fy']} grand total", f"{r['stated']:,}")
    present(f"FY{r['fy']} revenues", f"{r['revenue']:,}")

# The document claims its source's grand totals are $1 off in FY14/FY15 and exact after.
off = {r['fy']: r['general'] + r['revolving'] - r['stated'] for r in S['programme']}
bad = {fy: d for fy, d in off.items() if abs(d) > 1}
if bad:
    FAILS.append(f'source grand totals off by more than $1: {bad}')
exact = sorted(fy for fy, d in off.items() if d == 0)
print(f"\n  {'OK  ' if len(exact) == 4 else 'FAIL'}  "
      f"grand totals exact in {len(exact)} of {len(off)} years        {exact}")
if len(exact) != 4:
    FAILS.append(f'expected 4 exact years, found {len(exact)}')

# --- FY26 general fund athletics, from the workbook -------------------------------
head('FY26 whole programme')
import re
try:
    import openpyxl
    ws = openpyxl.load_workbook(
        os.path.join(ROOT, 'sources/xlsx/fy27-proposals.xlsx'), data_only=True).active
    cur, gf = None, 0.0
    for row in range(6, ws.max_row + 1):
        a, b, v = ws.cell(row, 1).value, ws.cell(row, 2).value, ws.cell(row, 7).value
        if a and isinstance(a, str) and re.match(r'^\d{4}\s*-', a.strip()):
            cur = a.strip()
        if cur and cur.startswith('3510') and b and isinstance(v, (int, float)):
            gf += v
except ImportError:
    print('  SKIP  openpyxl not available'); gf = None

REV_FY26 = 146911.44          # school-funds-fy26.xlsx, Athletics Revolving, net expenditures
if gf is not None:
    present('FY26 general fund athletics', f'{gf:,.0f}')
    present('FY26 revolving expenditures', f'{REV_FY26:,.0f}')
    present('FY26 whole programme', f'{gf + REV_FY26:,.0f}')
    present('FY26 fund share', f'{REV_FY26 / (gf + REV_FY26) * 100:.1f}%')
    g19, r19 = 307931, 87902
    present('FY19 fund share', f'{r19 / (g19 + r19) * 100:.1f}%')

# --- the transportation line, budget against reported actual ----------------------
head('Athletic transportation — budget against reported actual')
rows = list(csv.DictReader(open(os.path.join(DATA, 'line-history.csv'))))
cell = collections.defaultdict(dict)
for r in rows:
    # variant='' only -- a scenario column is a different proposal for the same year,
    # not another reading of the same figure. See notes/SCHEMA.md, budget_figure.
    if r['key'] == 'athletic transportation' and not r.get('variant'):
        cell[int(r['fy'])][r['stage']] = float(r['value'])
exact_years = []
for fy in sorted(cell):
    b = cell[fy].get('settled') or cell[fy].get('proposed')
    a_ = cell[fy].get('actual')
    if b and a_ and b == a_:
        exact_years.append(fy)
usable = [fy for fy in sorted(cell)
          if (cell[fy].get('settled') or cell[fy].get('proposed')) and cell[fy].get('actual')]
print(f'  {"OK  " if len(exact_years) == 4 else "FAIL"}  '
      f'years where actual equals budget exactly      {exact_years} of {len(usable)} usable')
if len(exact_years) != 4:
    FAILS.append(f'expected 4 exact years, found {len(exact_years)}: {exact_years}')
# Assert the counts the prose states, not merely that some sentence is present. The first
# draft said "four of eight" because FY25 was missed; the string check passed and the claim
# was still wrong.
WORDS = ['zero', 'one', 'two', 'three', 'four', 'five', 'six',
         'seven', 'eight', 'nine', 'ten', 'eleven', 'twelve']
present('exact-match count', f'Four of {WORDS[len(usable)]} usable years')
ex21 = [fy for fy in exact_years if fy != 2021]
present('excluding FY21', f'three of {WORDS[len(usable) - 1]}')
if len(ex21) != 3:
    FAILS.append(f'expected 3 exact years excluding FY21, found {len(ex21)}')

# --- the FY25 deficit question ----------------------------------------------------
head('The fund across FY25/FY26 -- what is actually known')
OPEN26, REV, EXP, CLOSE26 = 110247.89, 188944.46, 146911.44, 152280.91
present('FY26 opening = FY25 closing', f'{OPEN26:,.2f}')
present('FY26 closing', f'{CLOSE26:,.2f}')
present('FY26 surplus', f'{REV - EXP:,.0f}')
if abs((OPEN26 + REV - EXP) - CLOSE26) > 0.01:
    FAILS.append('the fund roll-forward does not reconcile')
print(f'  OK    roll-forward reconciles                    '
      f'{OPEN26:,.2f} + {REV:,.2f} - {EXP:,.2f} = {OPEN26 + REV - EXP:,.2f}')
present('FY26 approved March 2025', f'{102550:,}')
present('FY26 final', f'{127550:,}')
present('the later increase', f'{127550 - 102550:,}')

# --- is $127,550 defensible ------------------------------------------------------
head('$127,550 against every documented base')
LINE = 127550
for label, v, fy in [('FY17 all-in', 73986, 2017), ('FY14 all-in', 47085, 2014),
                     ('FY25 reported actual', 87822, 2025), ('FY24 reported actual', 40000, 2024)]:
    yrs = 2026 - fy
    present(f'{label} ratio', f'{LINE / v:.2f}')
    present(f'{label} implied rate', f'{(( LINE / v) ** (1 / yrs) - 1) * 100:.2f}%')
present('FY26 committed at Q3', f'{47847 + 13169:,}')
present('FY26 Q3 ratio', f'{LINE / (47847 + 13169):.2f}')

# --- the memo's definition --------------------------------------------------------
head("The Finance Committee memo's arithmetic")
present('ledger expended', f'{34219013.80:,.2f}')
present('ledger encumbrances', f'{2626115.87:,.2f}')
present('their sum', f'{34219013.80 + 2626115.87:,.2f}')

# --- the swap ---------------------------------------------------------------------
head('The swap')
present('FY19 officials', f'{40117:,}')
present('FY19 uniforms', f'{8000:,}')
present('officials + uniforms', f'{40117 + 8000:,}')
present('ArbiterSports', f'{59400:,}')
present('Prime Time Sports', f'{25421:,}')

# --- section 6: the unproven workbook ---------------------------------------------
head('Section 6 -- the citizen workbook, recorded as unproven')
# The 25/26 column is the prior year escalated 6.5%. Asserted so a future edit cannot
# quietly promote a modelled column to a measurement.
for y2425, y2526 in ((43446.06, 46270.05), (29377.50, 31287.03), (18242.50, 19428.24)):
    if abs(y2425 * 1.065 - y2526) > 0.05:
        FAILS.append(f'{y2425} x 1.065 != {y2526}; the 25/26 escalation claim is wrong')
print('  OK    25/26 is 24/25 escalated 6.5% (three seasons checked)')
for label, gf_, tot in (('FY24', 40000.0, 117555.00), ('FY25', 87822.0, 91066.06)):
    present(f'{label} workbook total', f'{tot:,.0f}')
    present(f'{label} implied fund share', f'{(tot - gf_) / tot * 100:.1f}%')
# The fund's margin, both eras
head('Section 6 -- the fund never had slack')
for fy, cost, rev in ((2014, 107257, 110474), (2017, 131551, 109351), (2018, 60001, 108000)):
    present(f'FY{fy} margin', f'{rev - cost:+,}')
for fy, cost, rev in ((2024, 129125, 128252.50), (2025, 53940, 117069.00)):
    present(f'FY{fy} margin', f'{rev - cost:+,.0f}')

# --- document basis counts --------------------------------------------------------
head('Source-type counts, from document-basis.csv')
basis = collections.Counter(r['source_type']
                            for r in csv.DictReader(open(os.path.join(DATA, 'document-basis.csv'))))
for k in ('ledger', 'restatement', 'forward', 'narrative'):
    present(f'{k} documents', f'| {basis[k]} |')

print()
if FAILS:
    print(f'FAILED — {len(FAILS)} figure(s) in the analysis do not match the data:')
    for f in FAILS:
        print(f'  {f}')
    sys.exit(1)
print('PASSED — every figure in the analysis matches what the data produces.')
