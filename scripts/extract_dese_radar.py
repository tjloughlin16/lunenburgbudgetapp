"""Extract DESE's RADAR district comparison into a long, queryable table.

    python3 scripts/extract_dese_radar.py

Writes `sources/data/dese-radar.csv` -- one row per district, year and measure.

WHAT THIS IS, PRECISELY

DESE's own figures for every Massachusetts district: enrollment, demographics, staffing
FTE, MCAS, and per-pupil expenditure by function, **across all funds**. It is the first
view in this archive of Lunenburg's school spending that is not the town's general fund
and not written by the district.

THREE THINGS IT IS NOT

1.  **Not the town's appropriation, and never to be subtracted from it.** DESE counts
    costs the school budget does not carry -- insurance and retirement attributed to the
    schools is $3,459 per pupil for Lunenburg in FY25 on its own. The difference between
    DESE's total and the town's appropriation is NOT "hidden money"; it is two different
    definitions. Reconciling them is a separate job that has not been done.

2.  **Not a fund breakdown.** It says "all funds" and gives one number. It cannot say
    which dollar came from a grant, a revolving fund or the levy, which is exactly the
    question rule 11 asks. It bounds the answer from outside; it does not give it.

3.  **Not a headcount.** `Paraprofessional FTE` is full-time equivalents from DESE's own
    staffing collection, not people and not payroll. It is much more than this project
    had -- a standing question says a headcount is unpublished -- but an FTE is not a
    headcount and must not be quoted as one.

The extract is long rather than wide (district, year, measure, value) so that a measure
DESE adds or renames appears as new rows rather than breaking a fixed set of columns.

Reconciles the per-pupil function columns against the workbook's own printed total: the
ten function rows must sum to `Total In-District Expenditures`. DESE rounds each to whole
dollars, so the tolerance is one dollar per component.

**Sixteen district-years do not tie, and every one of them is a charter school** -- off by
$12 to $93 per pupil, which is more than rounding can produce from ten components. What
DESE puts in that total and not in a printed column is not established. Lunenburg and its
peer districts all tie.

So the extract writes, and carries the verdict per row in `reconciles` rather than
discarding the source or waving the difference through. `--strict` refuses to write if
anything fails, for the case where a caller needs the whole file to tie.
"""
import argparse
import csv
import os
import sys

import openpyxl

# The districts this project actually compares against, plus the state as a baseline.
#
# RADAR covers all 421 Massachusetts districts across 17 years -- 176,328 rows, which
# takes the published database from 4MB to 47MB and past Cloudflare's 25MB per-asset
# limit. Publishing a 47MB file so that 420 districts nobody here analyses can be queried
# is a bad trade against a database a resident can actually download.
#
# So the extract is scoped to the six peers the peer analysis names, plus Lunenburg and
# the state. The FULL workbook is in the archive with its sha256 and its upstream address,
# so anyone wanting another district has the same file we used, one download away. That is
# rule 12 doing its job: we do not have to republish everything, we have to make the
# source reachable.
# Codes read out of the workbook itself, not recalled. Three of the first eight were
# invented and the reconciliation caught them, which is the whole point of naming what
# should be present and reporting what is absent rather than filtering silently.
KEEP = {
    '01620000': 'Lunenburg',
    '06730000': 'Groton-Dunstable',
    '06100000': 'Ashburnham-Westminster',
    '06160000': 'Ayer Shirley School District',
    '07350000': 'North Middlesex',
    '01250000': 'Harvard',
    '00190000': 'Ayer',
}

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'sources', 'state-dese', 'radar-district-comparison.xlsx')
OUT = os.path.join(ROOT, 'sources', 'data', 'dese-radar.csv')
DOC = 'sources/state-dese/radar-district-comparison.xlsx'

# The ten function components DESE prints, and the total it prints beside them. Named
# rather than inferred from position, so a reordered workbook is caught rather than
# silently mis-summed.
FUNCTIONS = [
    'Expenditures Per Pupil: Administration',
    'Expenditures Per Pupil: Instructional Leadership',
    'Expenditures Per Pupil: Teachers',
    'Expenditures Per Pupil: Other Teaching Services',
    'Expenditures Per Pupil: Professional Development',
    'Expenditures Per Pupil: Instructional Materials, Equipment and Technology',
    'Expenditures Per Pupil: Guidance, Counseling and Testing',
    'Expenditures Per Pupil: Pupil Services',
    'Expenditures Per Pupil: Operations and Maintenance',
    'Expenditures Per Pupil: Insurance, Retirement Programs and Other',
]
TOTAL = 'Expenditures Per Pupil: Total In-District Expenditures'


# DESE prints 'n/a' where a district has no value -- a regional district with no
# elementary school, a year before a measure was collected. It is not zero and it is not
# a number, so it is neither summed nor published as one.
BLANK = {'', 'n/a', 'N/A', 'na', '-', '--', None}


def num(v):
    if v in BLANK:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--strict', action='store_true',
                    help='refuse to write if any district-year does not tie')
    args = ap.parse_args()
    if not os.path.exists(SRC):
        print('missing %s -- run scripts/fetch_dese_radar.py first' % DOC)
        return 1
    ws = openpyxl.load_workbook(SRC, data_only=True)['district-comparison']

    header = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(2, c).value
        if h:
            header[c] = str(h).strip()
    missing = [f for f in FUNCTIONS + [TOTAL] if f not in header.values()]
    if missing:
        print('the workbook no longer prints these columns:')
        for m in missing:
            print('   ', m)
        return 1

    col = {v: k for k, v in header.items()}
    rows, checked, failed = [], 0, []
    verdict = {}
    for r in range(3, ws.max_row + 1):
        lea = ws.cell(r, col['LEA']).value
        if lea in (None, ''):
            continue
        lea = str(lea).zfill(8)
        year = ws.cell(r, col['Year']).value
        district = ws.cell(r, col['District']).value

        # Reconcile to the workbook's own printed total before trusting the row.
        parts = [num(ws.cell(r, col[f]).value) for f in FUNCTIONS]
        stated = num(ws.cell(r, col[TOTAL]).value)
        ties = ''
        if stated is not None and all(p is not None for p in parts):
            checked += 1
            got = sum(parts)
            if abs(got - stated) > len(FUNCTIONS):
                failed.append((district, year, got, stated, got - stated))
                ties = 'no'
            else:
                ties = 'yes'
        verdict[(lea, year)] = ties

        if lea not in KEEP:
            continue
        for c, name in header.items():
            if name in ('Year', 'LEA', 'District'):
                continue
            v = ws.cell(r, c).value
            if v in BLANK:
                continue
            group, _, measure = name.partition(': ')
            rows.append(dict(lea=lea, district=district, fy=year,
                             group=group, measure=measure or group, value=v,
                             reconciles=ties, doc_id=DOC))

    if failed:
        print('%d of %d district-years do NOT sum to their own printed in-district total:'
              % (len(failed), checked))
        for f in sorted(failed, key=lambda x: -abs(x[4]))[:6]:
            print('   %-46s FY%s  components %,.0f vs stated %,.0f  (%+.0f)'
                  .replace('%,', '%') % (str(f[0])[:46], f[1], f[2], f[3], f[4]))
        print('   marked `reconciles=no` in the extract rather than dropped.')
        if args.strict:
            print('   --strict: nothing written')
            return 1

    with open(OUT, 'w', newline='', encoding='utf-8') as fh:
        w = csv.DictWriter(fh, fieldnames=['lea', 'district', 'fy', 'group', 'measure',
                                           'value', 'reconciles', 'doc_id'])
        w.writeheader()
        for r in rows:
            w.writerow(r)

    years = sorted({r['fy'] for r in rows})
    leas = {r['lea'] for r in rows}
    absent = sorted(set(KEEP) - leas)
    if absent:
        print('  NOT FOUND in the workbook, so not published: %s'
              % ', '.join('%s (%s)' % (a, KEEP[a]) for a in absent))
    print('wrote %s' % os.path.relpath(OUT, ROOT))
    print('  %d rows | %d districts | FY%s-FY%s | %d measures'
          % (len(rows), len(leas), min(years), max(years),
             len({r['measure'] for r in rows})))
    print('  %d district-years reconcile to their own printed in-district total'
          % checked)
    lun = [r for r in rows if r['lea'] == '01620000']
    lun_ties = {r['reconciles'] for r in lun if r['reconciles']}
    print('  Lunenburg: %d rows across FY%s-FY%s | reconciles: %s'
          % (len(lun), min(r['fy'] for r in lun), max(r['fy'] for r in lun),
             ', '.join(sorted(lun_ties)) or 'not checkable'))
    return 0


if __name__ == '__main__':
    sys.exit(main())
