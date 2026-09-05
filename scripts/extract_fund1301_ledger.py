"""Extract the athletics revolving fund's cash journal from the town's MUNIS export.

Three workbooks arrived from the Town on 17 June 2026 in response to a public records
request. Each is a MUNIS "Journal Detail Export" for one fiscal year. Despite the request
asking for athletics financial data generally, every row in all three is the same single
account:

    D2 = '1301-0-000-0000-00-0-00-0-104000'   ORG 1301, OBJECT 104000, DESCRIPTION 'CASH'

So this is not the expense-object detail. It is the **cashbook** of fund 1301 — the
Chapter 658 athletics revolving fund — one row per receipt and per disbursement. What it
gives us that nothing else in the archive does is *dates* and a *running balance*: the
path through the year, not just the endpoint.

Rule 13 applies hard here. Two things in this file are ours and not the town's:

  * **The running balance.** MUNIS exports rows, not a balance. We compute one by sorting
    on EFF DATE. That ordering is a choice: many rows are backdated (a row effective
    2024-06-30 posted 2024-09-18), and rows sharing an effective date have no defined
    order. So the running balance is a reconstruction, and the column is named
    `running_balance_derived` to keep that visible in the data itself.
  * **Any reading of what a journal entry means.** The four largest movements in FY25 are
    source GEN with reference 'ADJ EXP' and a comment naming a memo we do not hold. We
    record the cells. We do not say what they were for.

What is *not* ours, and is the reason this file can be trusted at all: each year's export
carries its own opening balance as a row (SRC='SOY', REFERENCE='SOY BAL'). That gives the
extract a total the source itself prints, in the sense rule 13 requires — the closing
balance we compute for one year must equal the SOY row of the next. It does, to the cent,
across all three years. The script refuses to write if that chain ever breaks.

    python3 scripts/extract_fund1301_ledger.py

Writes sources/data/fund-1301-cash-journal.csv and prints the year-by-year roll-forward.
"""
import csv
import os
import sys
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'sources', 'town-ledgers', 'account-details')
OUT = os.path.join(ROOT, 'sources', 'data', 'fund-1301-cash-journal.csv')

# The publisher's own filenames, kept because links die and a resident asking the town for
# these has to ask by the name the town used (rule 12). Ours on the left, theirs on the
# right, in sources/town-ledgers/account-details/PROVENANCE-fund1301.md.
YEARS = [
    (2024, 'account-details-fy2024-fund1301.xlsx'),
    (2025, 'account-details-fy2025-fund1301.xlsx'),
    (2026, 'account-details-fy2026-fund1301.xlsx'),
]

SHEET = 'Journal Detail Export'

# Column letters, taken from row 1 of the export and asserted below rather than assumed.
HEADERS = ['ORG', 'OBJECT', 'PROJECT', 'ACCOUNT', 'DESCRIPTION', 'YEAR', 'PER', 'JOURNAL',
           'EFF DATE', 'POST DATE', 'SRC', 'T', 'REF1', 'PROJECT STRING', 'PO/REF2',
           'REF3', 'REFERENCE', 'AMOUNT', 'P', 'CHECK NO', 'WARRANT', 'VOUCHER',
           'CARRY FORWARD', 'VDR NAME/ITEM DESC', 'COMMENTS']

# MUNIS journal source codes seen in these three files. Descriptions are the standard
# MUNIS meanings and are *labels for reading*, not claims about any particular row.
SRC_MEANING = {
    'SOY': 'start of year — opening balance carried forward',
    'CRP': 'cash receipts posting',
    'APP': 'accounts payable — warrant disbursement',
    'PRJ': 'payroll journal',
    'GRV': 'general journal reversal / correction',
    'GEN': 'general journal entry',
}


def cell(row, i):
    v = row[i].value
    return '' if v is None else v


def load_year(fy, filename):
    path = os.path.join(SRC, filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f'{filename}: expected a sheet named {SHEET!r}, found {wb.sheetnames}')
    ws = wb[SHEET]

    got = [ws.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 1)]
    if got != HEADERS:
        sys.exit(f'{filename}: header row changed.\n  expected {HEADERS}\n  got      {got}')

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        org = cell(row, 0)
        if org == '':
            continue          # trailing blank row the export always writes
        if str(org) != '1301':
            sys.exit(f'{filename} row {i}: ORG {org!r}, expected 1301. '
                     'This extractor assumes the export is fund 1301 only.')
        acct = str(cell(row, 3))
        if not acct.endswith('104000'):
            sys.exit(f'{filename} row {i}: ACCOUNT {acct!r} is not the 104000 CASH object. '
                     'The export has widened beyond cash; this extractor would mis-total it.')
        rows.append(dict(
            fy=fy, source_row=i,
            period=cell(row, 6), journal=cell(row, 7),
            eff_date=cell(row, 8), post_date=cell(row, 9),
            src=cell(row, 10), ref1=cell(row, 12), po_ref2=cell(row, 14),
            ref3=cell(row, 15), reference=cell(row, 16),
            amount=float(cell(row, 17) or 0),
            check_no=cell(row, 19), warrant=cell(row, 20), voucher=cell(row, 21),
            vendor=cell(row, 23), comments=cell(row, 24),
        ))
    return rows


def d(v):
    return v.strftime('%Y-%m-%d') if isinstance(v, datetime) else ''


def main():
    years, chain = {}, []
    prior_close = None
    for fy, filename in YEARS:
        rows = load_year(fy, filename)
        soy = [r for r in rows if r['src'] == 'SOY']
        if len(soy) != 1:
            sys.exit(f'FY{fy}: expected exactly one SOY row, found {len(soy)}')
        opening = soy[0]['amount']
        txns = [r for r in rows if r['src'] != 'SOY']
        closing = opening + sum(r['amount'] for r in txns)

        # The reconciliation rule 13 asks for: last year's close is this year's printed
        # opening. Nothing here is checked against our own prose — it is checked against a
        # cell the town wrote.
        if prior_close is not None and round(prior_close, 2) != round(opening, 2):
            sys.exit(f'FY{fy}: SOY BAL {opening:,.2f} does not equal the FY{fy-1} closing '
                     f'balance {prior_close:,.2f} this script computed. The chain is broken; '
                     'refusing to write.')
        prior_close = closing

        years[fy] = rows
        chain.append(dict(fy=fy, opening=opening, closing=closing,
                          receipts=sum(r['amount'] for r in txns if r['amount'] > 0),
                          payments=sum(r['amount'] for r in txns if r['amount'] < 0),
                          n=len(txns)))

    # Running balance, ours and labelled as ours.
    out = []
    for fy, _ in YEARS:
        rows = years[fy]
        soy = [r for r in rows if r['src'] == 'SOY'][0]
        txns = sorted((r for r in rows if r['src'] != 'SOY'),
                      key=lambda r: (r['eff_date'] or datetime(1900, 1, 1), r['journal']))
        bal = soy['amount']
        out.append(dict(soy, running_balance_derived=bal))
        for r in txns:
            bal += r['amount']
            out.append(dict(r, running_balance_derived=round(bal, 2)))

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', newline='') as fh:
        w = csv.writer(fh)
        w.writerow(['fy', 'source_row', 'eff_date', 'post_date', 'src', 'src_meaning',
                    'period', 'journal', 'ref1', 'po_ref2', 'ref3', 'reference',
                    'amount', 'running_balance_derived', 'check_no', 'warrant', 'voucher',
                    'vendor', 'comments'])
        for r in out:
            w.writerow([r['fy'], r['source_row'], d(r['eff_date']), d(r['post_date']),
                        r['src'], SRC_MEANING.get(r['src'], ''), r['period'], r['journal'],
                        r['ref1'], r['po_ref2'], r['ref3'], r['reference'],
                        f"{r['amount']:.2f}", f"{r['running_balance_derived']:.2f}",
                        r['check_no'], r['warrant'], r['voucher'], r['vendor'],
                        r['comments']])

    print(f'wrote {OUT}  ({len(out)} rows)\n')
    print(f"{'':6} {'opening':>13} {'receipts':>13} {'payments':>13} {'net':>13} {'closing':>13}  txns")
    for c in chain:
        net = c['receipts'] + c['payments']
        print(f"FY{c['fy']}  {c['opening']:>13,.2f} {c['receipts']:>13,.2f} "
              f"{c['payments']:>13,.2f} {net:>13,.2f} {c['closing']:>13,.2f}  {c['n']:>4}")
    print('\nEach opening ties to the prior closing to the cent — the town printed the '
          'openings; we computed the closings.')

    # The rows that dominate the fund's year and are described only by a memo reference.
    print('\nGeneral-journal entries (SRC=GEN), raw cells, all three years:')
    for fy, _ in YEARS:
        for r in years[fy]:
            if r['src'] == 'GEN':
                print(f"  FY{fy} row {r['source_row']:>4}  eff {d(r['eff_date'])}  "
                      f"posted {d(r['post_date'])}  journal {r['journal']}  "
                      f"REF1={r['ref1']!r}  REFERENCE={r['reference']!r}  "
                      f"AMOUNT={r['amount']:>12,.2f}  COMMENTS={r['comments']!r}")


if __name__ == '__main__':
    main()
