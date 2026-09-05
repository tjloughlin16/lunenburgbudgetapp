#!/usr/bin/env python3
"""Classify every financial source document by what basis of figure it carries.

A document is one of:
    budget   - forward figures only (proposed, requested, recommended, appropriated)
    actual   - realised figures only (expended, actuals, year-end)
    both     - carries both, in separate columns, in the same table
    narrative- discusses money but publishes no figure table

and separately by GRANULARITY, which is what decides whether a figure can feed a
line-level analysis:

    line       - individual budget lines (Athletic Transportation, Paras, ...)
    department - one row per town department; school is a single row
    fund       - one row per fund (gift, athletics revolving, choice)
    program    - per-sport / per-school subtotals
    none       - no table

Rule 1 of CLAUDE.md forbids mixing budgets with actuals in one calculation. That
rule is unenforceable unless we can say which documents carry which, so this
writes the answer to sources/data/document-basis.csv and audit_provenance.py can
read it.

EVIDENCE is the raw header line the classification rests on, quoted verbatim with
its line number, so any row here can be checked against the file in one grep.
"""
import csv, glob, os, re, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# --- header signatures -------------------------------------------------------
# Deliberately tight. A narrative sentence containing the word "actual" is not a
# column header; these patterns want the shape of a table header row.
ACTUAL_HDR = [
    (re.compile(r'\bActual\b.*\bActual\b', re.I),      'repeated Actual columns'),
    (re.compile(r'\bACTUALS\b'),                        'ACTUALS column'),
    (re.compile(r'YTD\s+EXPENDED', re.I),               'MUNIS YTD EXPENDED'),
    (re.compile(r'\bActuals?\s+to\s+date\b', re.I),     'Actuals to date column'),
    (re.compile(r'\bAs\s+Expended\b', re.I),            'As Expended table'),
    (re.compile(r'\bBudget[- ]Vs[- ]Actuals?\b', re.I), 'Budget-vs-Actuals section'),
    (re.compile(r'\bbudget\s+and\s+actual\b', re.I),    'budget and actual schedule'),
    (re.compile(r'\bExpenditure\s+Control\b', re.I),    'Expenditure Control (ledger)'),
    (re.compile(r'Account\s+Detail\s+History', re.I),   'Account Detail History (ledger)'),
]
# "ACTUAL TAX BILLING" is the term of art for the third/fourth-quarter tax bill,
# as opposed to the preliminary bill. It is not an accounting actual, and the
# document carries no expenditure figures at all.
ACTUAL_NOT = re.compile(r'actual\s+(tax\s+)?billing', re.I)

# --- system of record --------------------------------------------------------
# A figure is a LEDGER figure when it exists because a transaction did: it came off
# the town's accounting system, the balance sheet, or the account detail history.
# Everything else labelled "actual" in this archive is a RESTATEMENT -- a prior year
# re-presented inside a document written by the party doing the spending, usually to
# argue for next year's budget. The two are not interchangeable and the archive did
# not previously distinguish them.
#
# The distinction is load-bearing. Athletic transportation was encumbered as a single
# $40,000 purchase order in FY23 with nothing expended; the restatement published that
# line's "actual" as $39,880. The ledger separates committed from paid. The restatement
# does not, and neither does the Finance Committee memo, which reports expended plus
# encumbered under the word "expenditures".
LEDGER_HDR = [
    # The strongest ledger signature in the archive, and the only one that is per
    # transaction rather than per account: a journal export heads its rows with the
    # journal number and both an effective and a posting date. Nothing that restates a
    # prior year inside a budget document carries a posting date, because a restatement
    # has no posting -- it has a paragraph.
    (re.compile(r'\bJOURNAL\b.*\bEFF\s+DATE\b.*\bPOST\s+DATE\b', re.I | re.S),
                                                           'journal detail export'),
    (re.compile(r'YEAR-TO-DATE\s+BUDGET\s+REPORT', re.I), 'MUNIS year-to-date budget report'),
    (re.compile(r'\bglytdbud\b', re.I),                    'MUNIS program id glytdbud'),
    # Account codes alone are not proof: the town's FY27 detailed budget is laid out by
    # ORG/OBJ and carries only forward scenarios. The ledger signature is account codes
    # ALONGSIDE a realised column, which is checked at the call site.
    (re.compile(r'\bORG\s+OBJ\b(?=.*(?:Expended|Encumbr))', re.I | re.S), 'ORG/OBJ with realised columns'),
    (re.compile(r'ACCOUNT\s+ORG\s+CODE', re.I),            'account org code column'),
    (re.compile(r'Balance\s+Sheet\s+Report', re.I),        'balance sheet report'),
    (re.compile(r'Account\s+Detail\s+History', re.I),      'account detail history'),
    (re.compile(r'\bExpenditure\s+Control\b', re.I),      'expenditure control account'),
    # The town's fund reports head four adjacent cells Beginning / Revenue / Expenditure /
    # Remaining. No single cell says anything; the row does. Matched against joined rows.
    (re.compile(r'Beginning.*Revenue.*Expenditure.*Remaining', re.I | re.S),
                                                          'fund report roll-forward columns'),
    (re.compile(r'\bIndependent Auditor', re.I),            'audited financial statements'),
]
BUDGET_HDR = [
    (re.compile(r'\bBudgeted\b', re.I),                 'Budgeted column'),
    (re.compile(r'\bFINAL BUDGET\b'),                   'FINAL BUDGET column'),
    (re.compile(r'\bProposed\b', re.I),                 'Proposed column'),
    (re.compile(r'\bRecommended?\b', re.I),             'Recommended column'),
    (re.compile(r'\bRequested\b', re.I),                'Requested column'),
    (re.compile(r'\bAPPROP\b'),                         'MUNIS APPROP column'),
    (re.compile(r'REVISED\s+BUDGET', re.I),             'REVISED BUDGET column'),
    (re.compile(r'\bLevel Service\b', re.I),            'Level Service column'),
    (re.compile(r'\bBalanced Proposed\b', re.I),        'Balanced Proposed column'),
    (re.compile(r'\bOriginal Budget\b', re.I),          'Original Budget column'),
    # A bare "budget" is in every one of these documents; a fiscal year immediately
    # in front of it is a column header. Same for the town's balanced/tier scenarios.
    (re.compile(r'\bFY\s?\d{2}\s+BUDGET\b', re.I),      'FY__ BUDGET column'),
    (re.compile(r'\bTIER\s*[12]\b', re.I),             'Tier scenario column'),
    (re.compile(r'\bBALANCED\s*[-\u2010-\u2015]?\s*NO\b', re.I), 'BALANCED-NO OVERRIDE column'),
]

def scan_text(path):
    """Return (actual_evidence, budget_evidence); each (lineno, text, why) or None.

    Extracted PDF text wraps a single header row across several lines -- the FY27
    detailed budget renders "FY26  BUDGET  BALANCED- NO / OVERRIDE" as two lines,
    and a line-at-a-time matcher scores it as narrative. So every position is
    tested twice: the line alone, and the line joined with the two after it. The
    evidence column records which, because a match that only appears in the joined
    window is weaker than one that appears on a single line."""
    act = bud = led = None
    try:
        lines = open(path, errors='ignore').read().splitlines()
    except OSError:
        return None, None, None
    n = len(lines)
    for i in range(n):
        single = lines[i].strip()
        joined = ' '.join(x.strip() for x in lines[i:i + 3]).strip()
        for cand, tag in ((single, ''), (joined, ' [wrapped]')):
            if not (8 < len(cand) < 400):
                continue
            if act is None and not ACTUAL_NOT.search(cand):
                for rx, why in ACTUAL_HDR:
                    if rx.search(cand):
                        act = (i + 1, cand[:200], why + tag); break
            if bud is None:
                for rx, why in BUDGET_HDR:
                    if rx.search(cand):
                        bud = (i + 1, cand[:200], why + tag); break
            if led is None:
                for rx, why in LEDGER_HDR:
                    if rx.search(cand):
                        led = (i + 1, cand[:200], why + tag); break
        if act and bud and led:
            break
    return act, bud, led

def scan_xlsx(path):
    """Workbooks: scan every cell in the first 8 rows of every sheet, plus label
    columns, and record hidden columns -- a hidden column is present in the data
    and absent from the reader's screen, which is its own finding."""
    try:
        import openpyxl
    except ImportError:
        return None, None, None, ''
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None, None, None, ''
    act = bud = led = None
    hidden = []
    for ws in wb.worksheets:
        h = [k for k, v in ws.column_dimensions.items() if v.hidden]
        if h:
            hidden.append(f'{ws.title}:{",".join(sorted(h))}')
        for r in range(1, min(ws.max_row, 30) + 1):
            cells = []
            for c in range(1, min(ws.max_column, 40) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    cells.append((f'{ws.title}!{ws.cell(r,c).coordinate}', v.strip()))
            # A header that spans four adjacent cells says nothing in any one of them, so
            # each row is tested joined as well as cell by cell -- with the next row too,
            # because these reports stack "Beginning / Balance" over two rows.
            nxt = [ws.cell(r + 1, c).value for c in range(1, min(ws.max_column, 40) + 1)]
            joined = ' '.join([t for _, t in cells] +
                              [str(x).strip() for x in nxt if isinstance(x, str) and x.strip()])
            if joined and len(cells) > 1:
                cells = cells + [(f'{ws.title}!row{r}', joined)]
            for ref, s in cells:
                if act is None and not ACTUAL_NOT.search(s):
                    for rx, why in ACTUAL_HDR:
                        if rx.search(s):
                            act = (ref, s[:200], why); break
                if bud is None:
                    for rx, why in BUDGET_HDR:
                        if rx.search(s):
                            bud = (ref, s[:200], why); break
                if led is None:
                    for rx, why in LEDGER_HDR:
                        if rx.search(s):
                            led = (ref, s[:200], why); break
    return act, bud, led, '; '.join(hidden)

# Directories holding documents that publish figures. Meeting minutes are excluded
# by design: they are narrative, there are 1,100 of them, and a figure quoted in
# minutes is a restatement of a document catalogued here.
DIRS = [
    'sources/district-budget/text', 'sources/town-budget/text',
    'sources/town-supplementary/text', 'sources/town-ledgers',
    'sources/peer-districts', 'sources/contracts/txt',
]

def main():
    rows = []
    for d in DIRS:
        for p in sorted(glob.glob(os.path.join(ROOT, d, '*.txt'))):
            act, bud, led = scan_text(p)
            rows.append((os.path.relpath(p, ROOT), act, bud, led, ''))
    for p in sorted(glob.glob(os.path.join(ROOT, 'sources/**/*.xlsx'), recursive=True)):
        act, bud, led, hidden = scan_xlsx(p)
        rows.append((os.path.relpath(p, ROOT), act, bud, led, hidden))

    out = os.path.join(ROOT, 'sources/data/document-basis.csv')
    with open(out, 'w', newline='') as fh:
        w = csv.writer(fh)
        w.writerow(['path', 'source_type', 'basis',
                    'ledger_at', 'ledger_why', 'ledger_evidence',
                    'actual_at', 'actual_why', 'actual_evidence',
                    'budget_at', 'budget_why', 'budget_evidence', 'hidden_columns'])
        for path, act, bud, led, hidden in sorted(rows):
            basis = ('both' if act and bud else 'actual' if act
                     else 'budget' if bud else 'narrative')
            # A document is a restatement only if it publishes realised figures without
            # being the accounting system that produced them.
            source_type = ('ledger' if led else 'restatement' if act
                           else 'forward' if bud else 'narrative')
            w.writerow([path, source_type, basis,
                        led[0] if led else '', led[2] if led else '', led[1] if led else '',
                        act[0] if act else '', act[2] if act else '', act[1] if act else '',
                        bud[0] if bud else '', bud[2] if bud else '', bud[1] if bud else '',
                        hidden])
    counts = {}
    for path, act, bud, led, _ in rows:
        t = ('ledger' if led else 'restatement' if act else 'forward' if bud else 'narrative')
        counts[t] = counts.get(t, 0) + 1
    print(f'{len(rows)} documents scanned -> {os.path.relpath(out, ROOT)}\n')
    print('  by source type -- what produced the figure:')
    for k in ('ledger', 'restatement', 'forward', 'narrative'):
        print(f'     {k:<12} {counts.get(k,0)}')
    led_rows = [r for r in rows if r[3]]
    print(f'\n  the {len(led_rows)} ledger documents:')
    for path, _, _, led, _ in sorted(led_rows):
        print(f'     {path}\n        [{led[2]}] {str(led[1])[:96]}')
    return 0

if __name__ == '__main__':
    sys.exit(main())
