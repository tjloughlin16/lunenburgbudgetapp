"""Extract the district's sport-by-sport athletics workbook into long form.

`athletics-by-sport-fy2024-fy2026.xlsx` came from the Town on 17 June 2026 by records request.
The publisher's filename is `Copy of Athletics 24.25 (1).xlsx`. It is the only document in
the archive that puts participation counts and cost lines side by side for one sport in one
year, and it covers three school years — 23/24, 24/25 and 25/26 — on every sheet.

It is also, as far as we can tell, the document underneath `Athletics_v10.xlsx`, the
resident's analysis recorded in `sources/analyses/athletics.md` §6 as UNPROVEN. Holding
this changes the status of the figures that rested on it: they now rest on a workbook the
Town supplied, not on an unsourced copy.

**What this script does and does not assert.**

The workbook's three sheets do not share a column layout — `Official` starts at column
`AF` on Fall and at `AG` on Winter and Spring, and every column after it is shifted by one.
Hard-coding either layout would silently read Assignor as Official on two sheets out of
three. So the column map is built from the file: row 1 carries the metric name over the
first of a three-column block, row 2 carries the school year in each column of it. Nothing
about position is assumed.

The sheets carry their own `Total Hs`, `Total MS` and `Total` rows. Rule 13 says an extract
with a total the source prints must reconcile to it, so this script does, for every metric
and every year, and writes the result to a second file. It does **not** refuse to write on a
mismatch, and the reason is worth stating: several of those totals are wrong in the source —
blank where the rows beneath them have values, or stale. That is a property of the document,
not a defect in reading it, and burying it would be the error. The mismatches are published
as data. Where a figure in an analysis rests on a column whose total does not tie, the
analysis has to say so.

Counts are sometimes written with an annotation rather than as a number — `'1 (130) 2 ( 140)'`
in one cell, `'1 (150)'` in another. Those are kept verbatim in `raw` and parsed to a leading
integer in `value` only when that is unambiguous; otherwise `value` is empty and
`is_numeric` is 0.

    python3 scripts/extract_athletics_by_sport.py

Writes sources/data/athletics-by-sport.csv and
       sources/data/athletics-by-sport-reconciliation.csv
"""
import csv
import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BOOK = os.path.join(ROOT, 'sources', 'town-ledgers', 'account-details',
                    'athletics-by-sport-fy2024-fy2026.xlsx')
OUT = os.path.join(ROOT, 'sources', 'data', 'athletics-by-sport.csv')
RECON = os.path.join(ROOT, 'sources', 'data', 'athletics-by-sport-reconciliation.csv')

SEASONS = ['Fall', 'Winter', 'Spring']

# Row 2 spells the school year several ways. Normalised to the fiscal year it ends in,
# which is how every other file in sources/data/ is keyed: 24/25 -> FY2025.
YEAR_ALIASES = {
    '23/24': 2024, '2023-2024': 2024,
    '24/25': 2025, '2024-2025': 2025,
    '25/26': 2026, '25-/26': 2026, '2025-2026': 2026,
}

# Which metrics are money and which are people. Kept explicit because the reconciliation
# tolerance differs and because a reader of the CSV should not have to guess.
MONEY = {'Expected Fees Received', 'Due', 'Total Received', 'Fee Received',
         'Total Fees Received', 'Official', 'Assignor', 'Police/ EMS', 'Coaches',
         'Transportation', 'Equipment Recon', 'Costs for all 3 Seasons', 'Dues & Fees',
         'Uniforms', 'Equipment', 'Misc', 'Total Expenses'}
COUNTS = {'Total Athletes', 'Full Pay', '2nd Sibling', '3rd sibling', 'Full Waiver',
          'HS Red Fee', 'MS Red Fee'}

# The three "Total ..." rows each sheet prints, and what they are meant to total.
TOTAL_ROWS = {'total hs': 'HS', 'total ms': 'MS', 'total': 'ALL'}

# Rows in column A that are section banners carrying the fee in their text, not sports.
BANNER = re.compile(r'^(HS|MS)\b', re.I)


def norm(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip()


def build_column_map(ws):
    """(metric, fy) -> column index, read from rows 1 and 2 rather than assumed."""
    metric_at = {}
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(row=1, column=c).value)
        if v:
            metric_at[c] = v
    cols, current = {}, None
    for c in range(1, ws.max_column + 1):
        if c in metric_at:
            current = metric_at[c]
        y = norm(ws.cell(row=2, column=c).value)
        if not y or current is None:
            continue
        if y not in YEAR_ALIASES:
            sys.exit(f'{ws.title}!{get_column_letter(c)}2 = {y!r}: unrecognised school year')
        cols[(current, YEAR_ALIASES[y])] = c
    return cols


def parse(v):
    """Return (value_or_None, raw_string, is_numeric)."""
    if v is None or v == '':
        return None, '', True
    if isinstance(v, (int, float)):
        return float(v), repr(v), True
    raw = norm(v)
    m = re.match(r'^(-?\d+(?:\.\d+)?)\b', raw)
    # '1 (130) 2 ( 140)' means two different things at two prices; a leading 1 would be a
    # lie. Only take the number when nothing else numeric follows it.
    if m and not re.search(r'\d', raw[m.end():]):
        return float(m.group(1)), raw, True
    return None, raw, False


def read_sheet(ws):
    cols = build_column_map(ws)
    rows, level, seen_total = [], None, False
    totals = {}
    for r in range(3, ws.max_row + 1):
        name = norm(ws.cell(row=r, column=1).value)
        if not name:
            continue
        key = name.lower().rstrip('.').strip()
        if key in TOTAL_ROWS:
            scope = TOTAL_ROWS[key]
            for (metric, fy), c in cols.items():
                val, raw, ok = parse(ws.cell(row=r, column=c).value)
                totals[(scope, metric, fy)] = (val, raw, ok, r, get_column_letter(c))
            if scope == 'ALL':
                seen_total = True
            continue
        if seen_total:
            break                      # everything below the grand total is commentary
        # A section banner — 'HS -$250', 'MS- $200.' — announces a level and carries the
        # fee in its own text. It is told apart from a sport whose name starts the same way
        # ('HS Basketball') by having no athlete count in any year: every real sport row
        # has one. Testing for an empty row instead does not work, because the banner rows
        # sit alongside the shared-cost legend parked in the 'Costs for all 3 Seasons'
        # columns and are therefore not empty.
        if BANNER.match(key) and not any(
                parse(ws.cell(row=r, column=c).value)[0]
                for (m, _), c in cols.items() if m == 'Total Athletes'):
            level = 'HS' if key.upper().startswith('HS') else 'MS'
            continue
        if level is None:
            continue                   # row 3, the fee schedule strip
        for (metric, fy), c in cols.items():
            val, raw, ok = parse(ws.cell(row=r, column=c).value)
            if val is None and not raw:
                continue
            rows.append(dict(season=ws.title, level=level, sport=name, fy=fy,
                             metric=metric, value=val, raw=raw, is_numeric=int(ok),
                             cell=f'{get_column_letter(c)}{r}'))
    return rows, totals, cols


def main():
    if not os.path.exists(BOOK):
        sys.exit(f'missing {BOOK}')
    wb = openpyxl.load_workbook(BOOK, data_only=True)
    for s in SEASONS:
        if s not in wb.sheetnames:
            sys.exit(f'expected a sheet named {s!r}; found {wb.sheetnames}')

    all_rows, recon = [], []
    for s in SEASONS:
        ws = wb[s]
        rows, totals, cols = read_sheet(ws)
        all_rows += rows

        for (scope, metric, fy), (printed, raw, ok, trow, tcol) in sorted(totals.items()):
            ours = sum(r['value'] or 0 for r in rows if r['metric'] == metric
                       and r['fy'] == fy and (scope == 'ALL' or r['level'] == scope))
            has_rows = any(r['metric'] == metric and r['fy'] == fy
                           and (scope == 'ALL' or r['level'] == scope) for r in rows)
            if not has_rows and printed in (None, 0):
                continue
            diff = (printed or 0) - ours
            recon.append(dict(season=s, scope=scope, metric=metric, fy=fy,
                              printed=printed, printed_cell=f'{tcol}{trow}',
                              printed_raw=raw, summed_from_rows=round(ours, 2),
                              difference=round(diff, 2),
                              ties=int(abs(diff) < 0.005)))

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=['season', 'level', 'sport', 'fy', 'metric',
                                           'value', 'raw', 'is_numeric', 'cell'])
        w.writeheader()
        for r in sorted(all_rows, key=lambda r: (SEASONS.index(r['season']), r['level'],
                                                 r['sport'], r['fy'], r['metric'])):
            w.writerow(r)

    with open(RECON, 'w', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=['season', 'scope', 'metric', 'fy', 'printed',
                                           'printed_cell', 'printed_raw',
                                           'summed_from_rows', 'difference', 'ties'])
        w.writeheader()
        for r in recon:
            w.writerow(r)

    ties = sum(r['ties'] for r in recon)
    print(f'wrote {OUT}   ({len(all_rows)} rows)')
    print(f'wrote {RECON} ({len(recon)} checks, {ties} tie, {len(recon)-ties} do not)\n')

    print('Totals the workbook prints that its own rows do not add up to:')
    for r in recon:
        if not r['ties']:
            p = 'blank' if r['printed'] is None else f"{r['printed']:,.2f}"
            print(f"  {r['season']:<7} {r['scope']:<3} {r['metric']:<24} FY{r['fy']}  "
                  f"printed {p:>14} at {r['printed_cell']:<5}  rows sum to "
                  f"{r['summed_from_rows']:>13,.2f}  diff {r['difference']:>13,.2f}")

    print('\nTransportation, summed from the sport rows (not from the printed totals):')
    for fy in (2024, 2025, 2026):
        tot = 0
        parts = []
        for s in SEASONS:
            v = sum(r['value'] or 0 for r in all_rows
                    if r['season'] == s and r['metric'] == 'Transportation' and r['fy'] == fy)
            parts.append(f'{s} {v:,.2f}')
            tot += v
        print(f'  FY{fy}  ' + '  '.join(f'{p:>22}' for p in parts) + f'   total {tot:>12,.2f}')


if __name__ == '__main__':
    main()
