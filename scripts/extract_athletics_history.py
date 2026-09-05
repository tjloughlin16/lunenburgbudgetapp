#!/usr/bin/env python3
"""Athletics, both sides of the money, every year we can see either one.

Almost every figure in this project is a general fund appropriation, because that is what
the district publishes. Athletics is the exception: for FY14-FY19 the district published
the appropriation and the Chapter 658 revolving fund side by side, line by line, in
`fy19-proposed-athletics-budget.pdf`. For FY26 a records request produced the fund's own
year-end reconciliation. Between those two windows the fund is invisible.

This writes both sides into one series so the gap is legible rather than implied:

    sources/data/athletics-history.csv
        fy, side, item, amount, basis, source

`side` is `general` or `revolving`. `basis` is `actual`, `budget` or `unproven`.

The FY14-FY19 reconstruction is CHECKED, not asserted: the itemised rows are summed and
compared against the document's own stated column totals. Five of six years tie to the
dollar on each side; the exceptions are the document's own rounding and are printed.

    python3 scripts/extract_athletics_history.py
"""
import os, csv, sys, collections

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'sources/data/athletics-history.csv')
FY19_DOC = 'district-budget/docs/fy19-proposed-athletics-budget.pdf'
FUND_DOC = 'budget-workbooks/school-funds-fy26.xlsx'
WORKBOOK = 'Athletics_v10.xlsx (citizen analysis, unpublished)'

YEARS = [2014, 2015, 2016, 2017, 2018, 2019]

# --- FY14-FY19, transcribed from the FY19 document -------------------------------------
# Line numbers are of the text extraction, sources/district-budget/text/.
# Column headers are on lines 4 and 5: FY14..FY19 over Actual/Actual/Actual/Actual/
# Budgeted/Requested. So FY14-FY17 are actuals; FY18 is a budget; FY19 is a request.
GENERAL = {                                       # line
    'Athletic Coaches':                 [114636, 130661, 120646, 130020, 163146, 150636],   # 10
    'Athletic Officials':               [39875, 40469, 38997, 38059, 39221, 40117],         # 20
    'Athletic Transportation':          [17000, 21600, 23000, 23000, 33500, 24975],         # 16
    'Athletic Director':                [0, 0, 0, 0, 48000, 49440],                         # 12
    'Athletic Dues & Fees':             [9519, 9759, 9751, 10782, 12499, 12529],            # 19
    'Athletic Insurance':               [5990, 5990, 5990, 5990, 5990, 5990],               # 18
    'Athletic Replacement of Uniforms': [3787, 4960, 1669, 7944, 7872, 8000],               # 24
    'Athletic New Equipment':           [0, 1440, 5456, 5151, 5850, 6400],                  # 25
    'Athletic Equipment/Reconditioning':[3065, 4361, 5265, 157, 4500, 4500],                # 23
    'Special Detail/Athletic Events':   [444, 5559, 139, 1592, 5900, 5344],                 # 8
}
# The document's rows carrying "(658)" -- the revolving fund. Three rows carry no fund
# marker but reconcile only against the fund column, and are placed there on that
# arithmetic: New Equipment (line 26), Track and Field Payment (28), turf (27).
REVOLVING = {
    'Athletic Expenses/Supplies':       [31894, 38730, 32623, 37156, 27114, 28309],         # 22
    'Athletic Transportation':          [30085, 40742, 33308, 50986, 27450, 40000],         # 17
    'Athletic Coaches':                 [45278, 27221, 40092, 34815, 0, 0],                 # 11
    'Athletic Secretary':               [0, 5245, 7796, 8594, 5437, 5593],                  # 13
    'Track and Field Payment':          [0, 0, 0, 25000, 40000, 0],                         # 28
    'Field upgrades / turf':            [0, 0, 0, 0, 0, 10000],                             # 27
    'Athletic New Equipment':           [890, 0, 0, 0, 3000, 4000],                         # 26
}
# The document's own stated column totals, lines 30 and 31.
STATED_GENERAL   = [194316, 224799, 210911, 222695, 326478, 307931]
STATED_REVOLVING = [108147, 111938, 113819, 156550, 103001, 87902]
# Line 38: Total Revenues (Fees & Gates); line 35/36 split it.
REVENUE = {'Fees': [None, 121875, 103660, 95360, 90000, 90000],
           'Gate receipts': [None, 13892, 17796, 13191, 18000, 18000]}
STATED_REVENUE = [110474, 140748, 121555, 109351, 108000, 108000]

BASIS_FY19 = {2014: 'actual', 2015: 'actual', 2016: 'actual', 2017: 'actual',
              2018: 'budget', 2019: 'budget'}


def check(name, items, stated):
    bad = []
    for i, fy in enumerate(YEARS):
        got = sum(v[i] for v in items.values())
        if got != stated[i]:
            bad.append((fy, got, stated[i], got - stated[i]))
    exact = len(YEARS) - len(bad)
    print(f'  {name:<22} {exact} of {len(YEARS)} years tie to the dollar')
    for fy, got, st, d in bad:
        print(f'      FY{fy}: itemised {got:,} vs the document\'s stated {st:,}  ({d:+,})')
    return bad


def main():
    rows = []
    print('Reconstructing the FY19 document from its own line items:')
    bad_g = check('general appropriation', GENERAL, STATED_GENERAL)
    bad_r = check('revolving fund 658', REVOLVING, STATED_REVOLVING)
    # A larger miss than the document's own rounding means the transcription is wrong.
    for fy, got, st, d in bad_g + bad_r:
        if abs(d) > 2:
            print(f'\nFAILED: FY{fy} is off by {d:+,}, which is more than this document\'s '
                  f'own rounding. The transcription is wrong.')
            return 1

    for i, fy in enumerate(YEARS):
        for side, items in (('general', GENERAL), ('revolving', REVOLVING)):
            for item, vals in items.items():
                if vals[i]:
                    rows.append(dict(fy=fy, side=side, item=item, amount=vals[i],
                                     basis=BASIS_FY19[fy], source=FY19_DOC))
        for item, vals in REVENUE.items():
            if vals[i]:
                rows.append(dict(fy=fy, side='revolving', item=f'REVENUE — {item}',
                                 amount=vals[i], basis=BASIS_FY19[fy], source=FY19_DOC))

    # --- FY20-FY25 general side, from the extracted line history -----------------------
    hist = list(csv.DictReader(open(os.path.join(ROOT, 'sources/data/line-history.csv'))))
    KEYS = {'athletic coaches': 'Athletic Coaches',
            'athletic officials': 'Athletic Officials',
            'athletic officials special detail coaches': 'Athletic Officials',
            'athletic transportation': 'Athletic Transportation',
            'athletic director': 'Athletic Director',
            'athletic trainer': 'Athletic Trainer',
            'athletic secretary': 'Athletic Secretary',
            'athletic dues fees': 'Athletic Dues & Fees',
            'athletic insurance': 'Athletic Insurance',
            'athletic replacement of uniforms': 'Athletic Replacement of Uniforms',
            'athletic new equipment': 'Athletic New Equipment',
            'athletic equipment reconditioning': 'Athletic Equipment/Reconditioning',
            'athletic expenses supplies': 'Athletic Expenses/Supplies',
            'special detail athletic events': 'Special Detail/Athletic Events',
            'unified sports track basketball coach': 'Unified Sports Coach',
            'freshman ms coaches': 'Freshman & MS Coaches'}
    seen = collections.defaultdict(dict)
    for r in hist:
        # variant='' only -- see notes/reference/SCHEMA.md, budget_figure.
        if (r['key'] in KEYS and r['stage'] == 'actual' and not r.get('variant')
                and 2020 <= int(r['fy']) <= 2025):
            v = float(r['value'])
            if v:
                seen[int(r['fy'])][KEYS[r['key']]] = (v, r['source'])
    for fy in sorted(seen):
        for item, (v, src) in sorted(seen[fy].items()):
            rows.append(dict(fy=fy, side='general', item=item, amount=round(v, 2),
                             basis='actual', source=src))

    # --- FY26 both sides ---------------------------------------------------------------
    # General: the FY26 FINAL BUDGET column of the FY27 workbook, via model/athletics.py
    # is not used here -- read the workbook directly so this file has one dependency less.
    try:
        import openpyxl, re
        ws = openpyxl.load_workbook(
            os.path.join(ROOT, 'sources/budget-workbooks/fy27-proposals.xlsx'), data_only=True).active
        cur = None
        for r in range(6, ws.max_row + 1):
            a, b, v = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 7).value
            if a and isinstance(a, str) and re.match(r'^\d{4}\s*-', a.strip()):
                cur = a.strip()
            if cur and cur.startswith('3510') and b and isinstance(v, (int, float)) and v:
                rows.append(dict(fy=2026, side='general', item=b.strip().title(),
                                 amount=round(v, 2), basis='budget',
                                 source='budget-workbooks/fy27-proposals.xlsx'))
    except ImportError:
        print('  openpyxl missing — FY26 general side skipped')

    # Revolving FY26: the fund's own year-end reconciliation.
    for item, amt in (('Salaries (4 revolving-fund staff)', 30513.84),
                      ('Purchase of service (officials, uniforms, transportation, '
                       'ice time, dues)', 113602.40),
                      ('General supplies', 2795.20),
                      ('REVENUE — High school user fees (net)', 162870.55),
                      ('REVENUE — Middle school user fees (net)', 26073.91)):
        rows.append(dict(fy=2026, side='revolving', item=item, amount=amt,
                         basis='actual', source=FUND_DOC))

    # --- FY24/FY25, the citizen workbook. UNPROVEN, and marked so ----------------------
    for fy, transport_total, officials_total in ((2024, 117555.00, 51570.00),
                                                 (2025, 91066.06, 50696.00)):
        gf = {2024: 40000.0, 2025: 87822.0}[fy]
        rows.append(dict(fy=fy, side='revolving', item='Athletic Transportation',
                         amount=round(transport_total - gf, 2), basis='unproven',
                         source=WORKBOOK))
        rows.append(dict(fy=fy, side='revolving', item='Athletic Officials',
                         amount=officials_total, basis='unproven', source=WORKBOOK))

    with open(OUT, 'w', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=['fy', 'side', 'item', 'amount', 'basis', 'source'])
        w.writeheader()
        w.writerows(sorted(rows, key=lambda r: (r['fy'], r['side'], r['item'])))
    print(f'\n{len(rows)} rows -> {os.path.relpath(OUT, ROOT)}')
    yrs = sorted({r['fy'] for r in rows})
    both = sorted({r['fy'] for r in rows if r['side'] == 'revolving'})
    print(f'  years covered           {yrs[0]}-{yrs[-1]}')
    print(f'  years with a fund side  {both}')
    print(f'  years general fund only {[y for y in yrs if y not in both]}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
