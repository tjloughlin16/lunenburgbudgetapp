"""Build the FY26 school budget line by line with the town ledger beside it, summed by
function code on both sides.

    python3 scripts/build_code_reconciliation_xlsx.py

Writes `sources/data/fy26-code-reconciliation.xlsx`.

WHY THIS FILE EXISTS

`build_gross_budget_xlsx.py` puts the ledger on its own sheet and writes `no crosswalk`
in the column that would have joined them, because when it was written no published
document mapped a budget line to an account. That has changed: the MUNIS `ACCOUNT` column
carries a function code in its fourth segment, and the district's workbook prints the same
code over each group. So the two structures can now be set side by side ONE CODE AT A
TIME, and this file is for checking that by eye.

THE DISTRICT'S WORKBOOK DOES NOT SUM BY CODE. It prints a code over a group and then
totals only at the section level, so nobody reading it can see that a code holds a
different amount than the town's books say. Every code here therefore carries its own two
sums and their difference.

HOW LINES ARE PAIRED, AND WHY THAT IS THE WEAK PART

Within a code, each budget line is paired to a ledger account of the SAME AMOUNT, to the
dollar. **Amount is not a key.** It shows that a figure of that size exists on both sides;
it never shows the two are the same line, and where a code holds two lines of equal value
the pairing between them is arbitrary. The `PAIRED BY` column says so on every row, and
anything that could not be paired is written in amber rather than left blank -- a blank
cell reads as zero and as agreement, and it is neither.

Read it as: the code sums are the finding; the row pairings are a reading aid.

A ROW THAT PAIRS IS NOT A ROW THAT AGREES

Pass 2 matches against the REVISED budget, which pairs athletic insurance -- the workbook
carries the $9,000 left after $20,000 was moved out, not the $29,000 appropriated. Pairing
it makes the correspondence visible and would, on its own, make a real $20,000
disagreement look settled. So `BUDGET vs APPROP` carries the gap between the workbook
figure and the appropriation on every paired row, flagged wherever it is not zero. One row
in the department is flagged.

`check_function_crosswalk.py` deliberately does NOT run pass 2, for the same reason from
the other side: its job is to fail on things that do not reconcile, and pairing on a
different column would hide one.
"""
import csv
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, 'sources', 'data')
OUT = os.path.join(DATA, 'fy26-code-reconciliation.xlsx')
# Two departments: 301 is SCHOOL NON-RECURRING EXPENSE and holds the $40,000 curriculum
# adoption account the district's budget book carries under function 2110.
DEPTS, FY = ('300', '301'), '2026'

INK, MUTED, RULE = '1B1B1B', '6B6B6B', 'D9D3C7'
GAP_FILL = PatternFill('solid', fgColor='FDF0D5')
GAP_INK = '8A5A00'
HEAD_FILL = PatternFill('solid', fgColor='1B1B1B')
GRP_FILL = PatternFill('solid', fgColor='EDE7DA')
SUM_FILL = PatternFill('solid', fgColor='E4EDF4')
BAD_FILL = PatternFill('solid', fgColor='F7DDDD')
MONEY = '#,##0;[Red]-#,##0'
BOX = Border(bottom=Side(style='thin', color=RULE))


def m(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def gap(cell, label):
    cell.value = label
    cell.fill = GAP_FILL
    cell.font = Font(size=8, italic=True, color=GAP_INK)
    cell.alignment = Alignment(horizontal='center')


def load():
    led = [r for r in csv.DictReader(open(os.path.join(DATA, 'munis-ledger.csv')))
           if r['dept'] in DEPTS and r['fy'] == FY and r['level'] == 'account'
           and r['account_type'] == 'expense']
    book = [r for r in csv.DictReader(open(os.path.join(DATA, 'lps-budget-lines.csv')))
            if r['kind'] == 'line']
    return led, book


def match(book_lines, accounts):
    """Greedy pairing by amount, within a dollar, in two passes.

    Pass 1 matches the workbook against the ledger's ORIGINAL APPROP. Pass 2 matches
    whatever is left against the REVISED budget, because the workbook does not state the
    original appropriation everywhere: athletic insurance is appropriated at $29,000 with
    $20,000 moved out of it, and the workbook carries the $9,000 that is left. A single
    pass on `original` left that account reading "no budget line of this amount" while
    its counterpart sat unpaired three rows below.

    Which pass caught a row is reported, because "matched the revised budget and not the
    appropriation" is a fact about that line, not a detail of the method.

    Returns [(book_line, account_or_None, how)] and the accounts nothing claimed.
    """
    pool = list(accounts)
    out = []
    for col, tag in (('original', 'amount (approp)'), ('revised', 'amount (revised)')):
        nxt = []
        src = book_lines if col == 'original' else [b for b, a, _ in out if a is None]
        for b in src:
            v = m(b['fy26_final'])
            j = None
            if abs(v) >= 0.5:
                j = next((k for k, a in enumerate(pool)
                          if abs(m(a[col]) - v) <= 1.0), None)
            nxt.append((b, pool.pop(j) if j is not None else None, tag))
        if col == 'original':
            out = nxt
        else:
            claimed = {id(b): (a, t) for b, a, t in nxt if a is not None}
            out = [(b, claimed[id(b)][0], claimed[id(b)][1]) if id(b) in claimed
                   else (b, a, how) for b, a, how in out]
    return out, pool


def main():
    led, book = load()
    A, B, label = {}, {}, {}
    for r in led:
        A.setdefault(r['function'], []).append(r)
    for r in book:
        g = re.match(r'\s*(\d{4})', r['function_group'] or '')
        c = g.group(1) if g else ''
        B.setdefault(c, []).append(r)
        lbl = re.sub(r'^\d{4}\s*-\s*', '', (r['function_group'] or '').strip()).strip()
        label.setdefault(c, [])
        if lbl and lbl not in label[c]:
            label[c].append(lbl)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FY26 by code'
    ws.sheet_view.showGridLines = False

    ws.cell(1, 1, 'LUNENBURG FY2026 SCHOOL BUDGET, BY FUNCTION CODE, WITH THE TOWN '
                  'LEDGER BESIDE IT').font = Font(bold=True, size=12, color=INK)
    ws.cell(2, 1, 'Budget columns are the district\'s FY27 projection workbook. Ledger '
                  'columns are the Town Accountant\'s MUNIS report for FY2026 period 12.'
            ).font = Font(size=9, color=MUTED)
    ws.cell(4, 1, 'A row that pairs is not a row that agrees: BUDGET vs APPROP carries '
                  'the gap between the workbook figure and the appropriation, and is '
                  'flagged wherever it is not zero.').font = Font(size=9, color=GAP_INK)
    ws.cell(3, 1, 'Lines are paired BY AMOUNT within a code, which is not a key: it shows '
                  'a figure of that size exists on both sides, never that the two are the '
                  'same line. The code sums are the finding; the row pairings are a '
                  'reading aid.').font = Font(size=9, color=GAP_INK)

    head = ['CODE', 'LINE ITEM (district)', 'THEIR ROW', 'BUDGET FY26',
            'LEDGER ACCOUNT', 'LEDGER NAME', 'ORIGINAL APPROP', 'TRANSFERS',
            'REVISED', 'YTD EXPENDED', 'BUDGET vs APPROP', 'PAIRED BY']
    hr = 6
    for i, h in enumerate(head, 1):
        c = ws.cell(hr, i, h)
        c.font = Font(bold=True, size=8, color='FFFFFF')
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[hr].height = 28

    r = hr + 1
    summary, offbase = [], []
    for code in sorted(set(A) | set(B), key=lambda c: (c == '', c)):
        lines, accts = B.get(code, []), A.get(code, [])
        pairs, spare = match(lines, accts)
        shown = '%s — %s' % (code or '(no code in the workbook)',
                             ' / '.join(label.get(code, [])) or 'no group heading')
        ws.cell(r, 1, shown).font = Font(bold=True, size=9, color=INK)
        for c in range(1, len(head) + 1):
            ws.cell(r, c).fill = GRP_FILL
        r += 1

        for b, a, how in pairs:
            ws.cell(r, 1, code or '—').font = Font(size=8, color=MUTED)
            ws.cell(r, 2, b['line_item'].strip()).font = Font(size=9)
            ws.cell(r, 3, int(b['row'])).font = Font(size=8, color=MUTED)
            v = m(b['fy26_final'])
            ws.cell(r, 4, v).number_format = MONEY
            if a is None:
                if abs(v) < 0.5:
                    for c in range(5, 13):
                        ws.cell(r, c, '—').font = Font(size=8, color=MUTED)
                        ws.cell(r, c).alignment = Alignment(horizontal='center')
                else:
                    for c in range(5, 11):
                        gap(ws.cell(r, c), 'no account of this amount' if c == 5 else '')
                    gap(ws.cell(r, 12), 'UNPAIRED')
            else:
                ws.cell(r, 5, a['account']).font = Font(size=8, name='Menlo')
                ws.cell(r, 6, a['name'].strip()).font = Font(size=9)
                for c, k in ((7, 'original'), (8, 'transfers'), (9, 'revised'),
                             (10, 'expended')):
                    ws.cell(r, c, m(a[k])).number_format = MONEY
                # Pairing a workbook figure against the REVISED budget makes the row
                # match while leaving a real disagreement at the appropriation. That is
                # the smoothing this file exists not to do, so the gap against ORIGINAL
                # APPROP is carried in its own column and flagged whatever the row
                # matched on.
                gapv = v - m(a['original'])
                if abs(gapv) > 1.0:
                    ws.cell(r, 11, gapv).number_format = MONEY
                    ws.cell(r, 11).font = Font(bold=True, size=9, color='B00000')
                    ws.cell(r, 11).fill = BAD_FILL
                    offbase.append((code, b['line_item'].strip(), a['name'].strip(),
                                    v, m(a['original']), gapv))
                else:
                    ws.cell(r, 11, '—').font = Font(size=8, color=MUTED)
                    ws.cell(r, 11).alignment = Alignment(horizontal='center')
                ws.cell(r, 12, how).font = Font(
                    size=8, color=GAP_INK if 'revised' in how else MUTED)
                ws.cell(r, 11).alignment = Alignment(horizontal='center')
            for c in range(1, len(head) + 1):
                ws.cell(r, c).border = BOX
            r += 1

        for a in spare:
            # An account appropriated NOTHING has no budget line to pair with, and that
            # is not a discrepancy -- it is a zero. But it is also not nothing: two of
            # these spent six figures. So they are separated from real unpaired rows and
            # the ones that spent are called out, because that is the loudest fact in
            # this file and it would otherwise sit in a list of harmless zeroes.
            # Three different states, and only one of them is a finding. An account
            # appropriated nothing that received a TRANSFER had a budget -- that is the
            # ordinary way a district covers something -- so it is not flagged. An
            # account that got neither and still paid out is the one worth seeing.
            zero = abs(m(a['original'])) < 0.5
            spent = abs(m(a['expended'])) >= 0.5
            moved = abs(m(a['transfers'])) >= 0.5
            naked = zero and spent and not moved
            gap(ws.cell(r, 2), ('NO APPROPRIATION AND NO TRANSFER — spent %s' % format(
                m(a['expended']), ',.0f')) if naked
                else 'budgeted by transfer only' if zero and spent
                else 'no appropriation' if zero
                else 'no budget line of this amount')
            ws.cell(r, 1, code or '—').font = Font(size=8, color=MUTED)
            ws.cell(r, 5, a['account']).font = Font(size=8, name='Menlo')
            ws.cell(r, 6, a['name'].strip()).font = Font(size=9)
            for c, k in ((7, 'original'), (8, 'transfers'), (9, 'revised'),
                         (10, 'expended')):
                ws.cell(r, c, m(a[k])).number_format = MONEY
            # This column answers ONE question -- did the row find a counterpart -- and
            # nothing else. It previously carried the account's budget state, which is a
            # different fact, and put 'budgeted by transfer only' where a reader was
            # looking for a reason the row did not pair. It is not one.
            gap(ws.cell(r, 12), 'UNPAIRED')
            if naked:
                for c in range(1, len(head) + 1):
                    ws.cell(r, c).fill = BAD_FILL
            for c in range(1, len(head) + 1):
                ws.cell(r, c).border = BOX
            r += 1

        bt = sum(m(x['fy26_final']) for x in lines)
        lt = sum(m(x['original']) for x in accts)
        et = sum(m(x['expended']) for x in accts)
        ws.cell(r, 2, 'TOTAL %s' % (code or 'uncoded')).font = Font(bold=True, size=9)
        ws.cell(r, 4, bt).number_format = MONEY
        ws.cell(r, 7, lt).number_format = MONEY
        ws.cell(r, 10, et).number_format = MONEY
        diff = lt - bt
        _bad = abs(diff) > max(1.0, len(accts))
        ws.cell(r, 12, 'differs by %s' % format(diff, ',.0f') if _bad else 'sums agree')
        ws.cell(r, 12).font = Font(bold=_bad, size=8,
                                   color='B00000' if _bad else MUTED)
        bad = abs(diff) > max(1.0, len(accts))
        for c in range(1, len(head) + 1):
            ws.cell(r, c).fill = BAD_FILL if bad else SUM_FILL
        for c in (4, 7, 10):
            ws.cell(r, c).font = Font(bold=True, size=9, color=INK)
            ws.cell(r, c).number_format = MONEY
        r += 2
        summary.append((code, ' / '.join(label.get(code, [])), len(lines), len(accts),
                        bt, lt, et, diff,
                        sum(1 for b, a, _ in pairs if a is None and abs(
                            m(b['fy26_final'])) >= 0.5)
                        + sum(1 for a in spare if abs(m(a['original'])) >= 0.5),
                        sum(1 for a in spare if abs(m(a['original'])) < 0.5
                            and abs(m(a['expended'])) >= 0.5
                            and abs(m(a['transfers'])) < 0.5)))

    for col, w in zip(range(1, 13),
                      (7, 46, 8, 14, 34, 13, 15, 12, 13, 14, 17, 22)):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(hr + 1, 5)

    s2 = wb.create_sheet('Code summary')
    s2.sheet_view.showGridLines = False
    s2.cell(1, 1, 'ONE ROW PER FUNCTION CODE — the sums the district\'s workbook does '
                  'not print').font = Font(bold=True, size=12, color=INK)
    h2 = ['CODE', 'WORKBOOK GROUP(S)', 'BUDGET LINES', 'LEDGER ACCOUNTS',
          'BUDGET FY26', 'LEDGER ORIGINAL', 'LEDGER EXPENDED', 'BUDGET vs LEDGER',
          'ROWS THAT DID NOT PAIR', 'SPENT WITH NO BUDGET AT ALL']
    for i, h in enumerate(h2, 1):
        c = s2.cell(3, i, h)
        c.font = Font(bold=True, size=8, color='FFFFFF')
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    s2.row_dimensions[3].height = 28
    rr = 4
    for code, lbl, nb, na, bt, lt, et, diff, unp, nb0 in summary:
        bad = abs(diff) > max(1.0, na)
        s2.cell(rr, 1, code or '(none)').font = Font(size=9, bold=True)
        s2.cell(rr, 2, lbl or 'no group heading').font = Font(size=9)
        s2.cell(rr, 3, nb).font = Font(size=9)
        s2.cell(rr, 4, na).font = Font(size=9)
        for c, v in ((5, bt), (6, lt), (7, et), (8, diff)):
            s2.cell(rr, c, v).number_format = MONEY
        s2.cell(rr, 9, unp).font = Font(size=9, bold=bool(unp),
                                        color='B00000' if unp else MUTED)
        s2.cell(rr, 10, nb0).font = Font(size=9, bold=bool(nb0),
                                         color='B00000' if nb0 else MUTED)
        if bad:
            for c in range(1, 11):
                s2.cell(rr, c).fill = BAD_FILL
        for c in range(1, 11):
            s2.cell(rr, c).border = BOX
        rr += 1
    s2.cell(rr, 2, 'TOTAL').font = Font(bold=True, size=9)
    for c, v in ((5, sum(x[4] for x in summary)), (6, sum(x[5] for x in summary)),
                 (7, sum(x[6] for x in summary)),
                 (8, sum(x[5] for x in summary) - sum(x[4] for x in summary))):
        s2.cell(rr, c, v).number_format = MONEY
        s2.cell(rr, c).font = Font(bold=True, size=9)
    for c in range(1, 11):
        s2.cell(rr, c).fill = SUM_FILL
    for col, w in zip(range(1, 11), (9, 52, 13, 15, 15, 16, 16, 16, 20, 22)):
        s2.column_dimensions[get_column_letter(col)].width = w
    s2.freeze_panes = s2.cell(4, 1)

    wb.save(OUT)
    nbad = sum(1 for x in summary if abs(x[7]) > max(1.0, x[3]))
    nunp = sum(x[8] for x in summary)
    nzs = sum(x[9] for x in summary)
    if offbase:
        print('  rows that pair but disagree with the appropriation:')
        for c, li, nm, v, o, g in offbase:
            print('    %-6s %-40s workbook %10s  approp %10s  %+10s'
                  % (c, li[:40], format(v, ',.0f'), format(o, ',.0f'),
                     format(g, ',.0f')))
    print('wrote %s' % os.path.relpath(OUT, ROOT))
    print('  %d codes; %d with sums that differ; %d rows that did not pair; '
          '%d account(s) spent with neither appropriation nor transfer'
          % (len(summary), nbad, nunp, nzs))
    return 0


if __name__ == '__main__':
    sys.exit(main())
