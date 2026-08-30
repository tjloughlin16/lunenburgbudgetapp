"""The DLS free cash proof for Lunenburg and eight comparable towns, 2021-2025.

Free cash is the certified, unrestricted amount a town may appropriate without raising
taxes — what is left at year end after everything committed is committed. It is the money
argued about when somebody says a town is hoarding, and the money pointed at when somebody
says a town is rebuilding. Two claims currently in the air about Lunenburg:

  * that the town is too conservative and is sitting on money it could spend, and
  * that free cash levels are "not up to standard", so the town is rebuilding.

**These files cannot settle that, and it is worth being exact about why.** A standard for
free cash is a RATIO — the Division of Local Services frames it against the operating
budget, and a town's own financial policy usually sets a target the same way. Every figure
in these workbooks is an absolute dollar amount with no denominator anywhere in the file.
Lunenburg at $3.35M and Shirley at $272K tells you nothing about which is closer to its
target until you know what each town spends. What the proof DOES support is each town's own
trend, and the composition of the number — where the money came from — which is a share and
therefore comparable.

What the file is, checked rather than assumed. One sheet, `A1` is the town name, `B3:F3`
are the years, and rows 4-17 are the proof. Two totals the source prints itself, both of
which this reconciles to before it will write:

  * rows 6-16 must sum to row 17, `Identified Free Cash July 1,`
  * row 5 `Current Year Calculation` in year N must equal row 4
    `Free Cash Certified Prior Year` in year N+1

Both tie to the dollar in all nine towns and all five years.

**A tenth file was supplied and is not included.** `FCPCompareAbington.xlsx` contains
Lunenburg's data — `A1` reads 'Lunenburg' and all 102 cells are identical to
`FCPCompareLunenburg.xlsx`. It is a mis-export, not an Abington profile, and treating it as
one would have put Lunenburg into the peer group twice and pulled every comparison toward
Lunenburg's own numbers. The filename was the only thing that said Abington.

    python3 scripts/extract_free_cash.py

Writes sources/data/free-cash-proof.csv
"""
import csv
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'sources', 'dls-free-cash')
OUT = os.path.join(ROOT, 'sources', 'data', 'free-cash-proof.csv')

TOWNS = ['ayer', 'groton', 'littleton', 'lunenburg', 'shirley', 'townsend', 'upton',
         'uxbridge', 'westford']

CERTIFIED = 'Current Year Calculation'
PRIOR = 'Free Cash Certified Prior Year'
IDENTIFIED = 'Identified Free Cash July 1,'


def read(town):
    path = os.path.join(SRC, f'free-cash-proof-{town}.xlsx')
    ws = openpyxl.load_workbook(path, data_only=True)['Sheet1']

    name = (ws['A1'].value or '').strip()
    if name.lower() != town:
        sys.exit(f'{os.path.basename(path)}: A1 says {name!r}, not {town!r}. The filename '
                 'is not the content — this is how a duplicate Lunenburg export arrived '
                 'labelled Abington.')

    years = []
    for c in range(2, 7):
        v = ws.cell(3, c).value
        if v is None:
            sys.exit(f'{town}: no year in row 3 column {c}')
        years.append(int(str(v).strip()))

    rows = {}
    for r in range(4, 18):
        label = (ws.cell(r, 1).value or '').strip()
        if not label:
            sys.exit(f'{town}: row {r} has no label')
        rows[label] = [ws.cell(r, 2 + i).value or 0 for i in range(5)]

    for key in (CERTIFIED, PRIOR, IDENTIFIED):
        if key not in rows:
            sys.exit(f'{town}: the proof has no row named {key!r}; the layout has changed')

    # Reconciliation 1: the components sum to the total the sheet prints.
    components = [k for k in rows if k not in (CERTIFIED, PRIOR, IDENTIFIED)]
    for i, y in enumerate(years):
        got = round(sum(rows[k][i] for k in components), 2)
        printed = round(rows[IDENTIFIED][i], 2)
        if got != printed:
            sys.exit(f'{town} {y}: components sum to {got:,.2f}, the sheet prints '
                     f'{printed:,.2f} for {IDENTIFIED!r}. Refusing to write.')

    # Reconciliation 2: this year's calculation is next year's prior-year certified.
    for i in range(len(years) - 1):
        if round(rows[CERTIFIED][i], 2) != round(rows[PRIOR][i + 1], 2):
            sys.exit(f'{town}: {years[i]} calculation {rows[CERTIFIED][i]:,.2f} does not '
                     f'equal {years[i+1]} prior-year certified {rows[PRIOR][i+1]:,.2f}')

    return name, years, rows, components


def main():
    out, checks = [], 0
    for town in TOWNS:
        name, years, rows, components = read(town)
        checks += len(years) + len(years) - 1
        for i, y in enumerate(years):
            for label, vals in rows.items():
                out.append(dict(
                    town=name, year=y, line=label, amount=f'{vals[i]:.2f}',
                    role=('certified' if label == CERTIFIED else
                          'prior_year_certified' if label == PRIOR else
                          'identified_total' if label == IDENTIFIED else 'component'),
                    source_file=f'dls-free-cash/free-cash-proof-{town}.xlsx',
                    source_ref=f'Sheet1!A{4 + list(rows).index(label)}'))

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=['town', 'year', 'line', 'amount', 'role',
                                           'source_file', 'source_ref'])
        w.writeheader()
        w.writerows(out)

    print(f'wrote {os.path.relpath(OUT, ROOT)}  ({len(out)} rows, {len(TOWNS)} towns)')
    print(f'  {checks} reconciliations against totals the source itself prints — all tie\n')

    by = {}
    for r in out:
        if r['role'] == 'certified':
            by.setdefault(r['town'], {})[r['year']] = float(r['amount'])
    years = sorted(next(iter(by.values())))
    print('  Certified free cash. ABSOLUTE DOLLARS — not comparable across towns of')
    print('  different size, because no denominator appears anywhere in these files.')
    print(f"  {'town':<11}" + ''.join(f'{y:>12}' for y in years))
    for t in sorted(by):
        print(f'  {t:<11}' + ''.join(f'{by[t][y]:>12,.0f}' for y in years))

    unex = 'Add Unencumbered/Unexpended Appropriations (CL#11)'
    share = {}
    for r in out:
        if r['line'] in (unex, IDENTIFIED):
            share.setdefault(r['town'], {}).setdefault(r['year'], {})[r['line']] = float(r['amount'])
    print('\n  Unspent appropriations as a share of identified free cash. A share HAS no')
    print('  size, so this one does compare between towns.')
    print(f"  {'town':<11}" + ''.join(f'{y:>11}' for y in years))
    for t in sorted(share):
        line = ''
        for y in years:
            d = share[t][y]
            line += f"{(d[unex] / d[IDENTIFIED] * 100 if d[IDENTIFIED] else 0):>10.0f}%"
        print(f'  {t:<11}{line}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
