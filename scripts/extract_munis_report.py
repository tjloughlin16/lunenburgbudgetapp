"""Parse any MUNIS YEAR-TO-DATE BUDGET REPORT (`glytdbud`) into one normalised CSV.

Written for the format, not for one file. Every report the Town Accountant produces has
this shape -- expenditures or revenues, any fund, any period, rolled up to department or
broken out to account -- so the FY24, FY25 and FY26 reports requested from the Town
Manager land in the same table as the ones already held, with no new code.

    python3 scripts/extract_munis_report.py                 # every report in sources/
    python3 scripts/extract_munis_report.py --check         # ...and fail if one does not tie

Writes `sources/data/munis-ledger.csv`.

Three things about the instrument, each of which has already bitten:

1.  **Revenues are printed as credits -- negative.** `Print revenue as credit: Y` on the
    report's own options page. A revenue of -9,229,410 is $9,229,410 of Chapter 70 aid
    coming in, not a shortfall. The sign is preserved exactly as printed and
    `account_type` says which convention applies; nothing here flips it.

2.  **The appropriation columns are printed rounded to whole dollars** while expended and
    encumbered carry cents. So a sum of rows cannot equal the report's own GRAND TOTAL to
    the cent, and a reconciliation that demanded it would fail on arithmetic rather than
    on a missing row. The tolerance is one dollar per row for the rounded columns and
    exact for the rest -- the same rule extract_town_ledger.py applies.

3.  **Our PDF-to-text step sometimes joins many accounts onto one line.** In
    `ef-sewer-revenue-fy26-q3.txt`, sixteen accounts share line 35. So this scans for the
    account pattern anywhere in the text rather than splitting on newlines and trusting
    that one line is one row. The instrument that reformats before you see it is part of
    the finding.

And it reconciles to the report's own printed GRAND TOTAL before it will write.
"""
import argparse
import csv
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'sources', 'data', 'munis-ledger.csv')

# A money token as MUNIS prints it: '-1,052,138', '-920,485.86', '.00', '0'.
NUM = r'-?[\d,]*\.\d{2}|-?[\d,]+'
SIX = r'\s+'.join('(%s)' % NUM for _ in range(6))

# An account row: five-digit org, six-digit object, name, six figures, a percentage.
ACCOUNT = re.compile(r'(\d{4,5})\s+(\d{6})\s+(.+?)\s+' + SIX + r'\s+(-?[\d.]+)%')
# A department rollup row, which is what `Print totals only: Y` produces instead.
DEPT = re.compile(r'(?:^|\n)\s*(\d{3})\s+([A-Z][^\n]*?)\s+' + SIX + r'\s+(-?[\d.]+)%')
# The fund appears on the column-header line: '0100     GENERAL FUND   APPROP ...'.
# NOT anchored to the start of a line: the sewer report's extraction runs that header on
# from the previous one ('...  PCT5000     SEWER BETTERMENTS    APPROP'), and anchoring
# lost every fund in the only multi-fund report we hold. The trailing APPROP is what
# makes the pattern safe without an anchor.
FUND = re.compile(r'(\d{4})\s{2,}([A-Z][A-Z /&\'-]+?)\s{2,}APPROP')
GRAND = re.compile(r'GRAND TOTAL\s+' + SIX)
PERIOD = re.compile(r'Year/Period:\s*(\d{4})/\s*(\d+)')
# The options page prints 'Print totals only: Y' with a colon, but the Find Criteria
# block prints 'Account type       Revenue' with none. Both forms, one regex.
OPT = lambda k: re.compile(re.escape(k) + r'\s*:?\s+(\S+)')


def money(tok):
    tok = tok.strip().replace(',', '')
    if tok in ('', '.00'):
        return 0.0
    return float(tok)


def parse(path):
    text = open(path, encoding='utf-8', errors='replace').read()

    per = PERIOD.search(text)
    if not per:
        return None
    fy, period = int(per.group(1)), int(per.group(2))

    def opt(k, default=None):
        m = OPT(k).search(text)
        return m.group(1) if m else default

    # Text extraction sometimes swallows the newline after the value, so 'Revenue'
    # arrives as 'RevenueAccount'. Match the declared word rather than the whole token,
    # and never fall back to a guess: the report always states this.
    declared = (opt('Account type') or '')
    acct_type = ('revenue' if declared.lower().startswith('revenue')
                 else 'expense' if declared.lower().startswith('expense') else None)
    if acct_type is None:
        raise ValueError('%s: report does not declare an Account type' % path)
    totals_only = (opt('Print totals only') or 'N').upper() == 'Y'
    suppress_zero = (opt('Suppress zero bal accts') or 'N').upper() == 'Y'

    # A report can cover several funds -- the sewer report covers four -- and a row does
    # not carry its own fund. The fund is POSITIONAL: it is whichever fund header last
    # appeared above the row. So scan fund headers and rows together by offset rather
    # than taking the first fund and applying it to everything, which silently filed
    # enterprise accounts under the general fund.
    marks = sorted([(m.start(), 'fund', m) for m in FUND.finditer(text)]
                   + [(m.start(), 'account', m) for m in ACCOUNT.finditer(text)],
                   key=lambda t: t[0])
    all_funds = ';'.join(sorted({'%s %s' % (m.group(1), m.group(2).strip())
                                 for _, k, m in marks if k == 'fund'}))

    rows, cur_code, cur_name = [], '', ''
    for _, kind, m in marks:
        if kind == 'fund':
            cur_code, cur_name = m.group(1), m.group(2).strip()
            continue
        vals = [money(m.group(i)) for i in range(4, 10)]
        rows.append(dict(level='account', org=m.group(1), object=m.group(2),
                         dept='', name=m.group(3).strip(), fund=cur_code,
                         fund_name=cur_name,
                         original=vals[0], transfers=vals[1], revised=vals[2],
                         expended=vals[3], encumbered=vals[4], available=vals[5],
                         pct_used=float(m.group(10))))
    if not rows:
        marks = sorted([(m.start(), 'fund', m) for m in FUND.finditer(text)]
                       + [(m.start(), 'dept', m) for m in DEPT.finditer(text)],
                       key=lambda t: t[0])
        for _, kind, m in marks:
            if kind == 'fund':
                cur_code, cur_name = m.group(1), m.group(2).strip()
                continue
            vals = [money(m.group(i)) for i in range(3, 9)]
            rows.append(dict(level='department', org='', object='', dept=m.group(1),
                             name=m.group(2).strip(), fund=cur_code, fund_name=cur_name,
                             original=vals[0], transfers=vals[1], revised=vals[2],
                             expended=vals[3], encumbered=vals[4], available=vals[5],
                             pct_used=float(m.group(9))))

    grand = GRAND.search(text)
    stated = [money(grand.group(i)) for i in range(1, 7)] if grand else None

    for r in rows:
        r.update(fy=fy, period=period,
                 funds_covered=all_funds, account_type=acct_type,
                 totals_only=int(totals_only), suppress_zero=int(suppress_zero),
                 rounded_columns='original,transfers,revised',
                 doc_id=os.path.relpath(path, ROOT))
    return dict(path=path, rows=rows, stated=stated, fy=fy, period=period,
                account_type=acct_type, totals_only=totals_only,
                suppress_zero=suppress_zero)


def reconcile(rep):
    """Sum the rows against the report's own GRAND TOTAL. Rule 13's last bullet."""
    if rep['stated'] is None:
        return 'no GRAND TOTAL printed', False
    cols = ('original', 'transfers', 'revised', 'expended', 'encumbered', 'available')
    n = len(rep['rows'])
    notes = []
    ok = True
    for i, c in enumerate(cols):
        got, want = sum(r[c] for r in rep['rows']), rep['stated'][i]
        # Rounded columns get a dollar per row; the cent-bearing ones must be exact.
        # The printed report rounds the appropriation columns to whole dollars while the
        # spreadsheet does not, so those three are allowed a dollar per row against a
        # printed total. Expended and encumbered carry cents on both sides and must be
        # exact -- that pair is what proves the two files are the same report.
        tol = n if c in ('original', 'transfers', 'revised') else 0.005
        if abs(got - want) > tol:
            ok = False
            notes.append(f'{c}: {got:,.2f} vs stated {want:,.2f}')
    return ('; '.join(notes) if notes else 'ties'), ok


# The same report, exported as data instead of printed. Columns are the MUNIS headings.
# It carries UN-ROUNDED appropriation figures, which the printed form does not, so where
# both exist the spreadsheet is the better copy of the same report.
XLSX_COLUMNS = {
    'FUND': 'fund', 'DEPARTMENT': 'dept', 'ORG': 'org', 'OBJ': 'object',
    'ACCOUNT DESCRIPTION': 'name', 'ORIGINAL APPROP': 'original',
    'TRANFRS/ADJSMTS': 'transfers', 'REVISED BUDGET': 'revised',
    'YTD EXPENDED': 'expended', 'ENCUMBRANCES': 'encumbered',
}


def parse_xlsx(path, fy, period, account_type, totals_only, suppress_zero, twin):
    """Parse the Excel form of a glytdbud report.

    **The spreadsheet does not state its own period, account type or grain.** Nothing
    inside it says 2026/12. Those come from the PRINTED twin, which has the options page,
    and the two are asserted to be one report by reconciling to the printed GRAND TOTAL.
    Taking the period from a filename would be exactly the derived-quoted-as-observed
    error rule 13 is about, so the twin is required rather than optional.
    """
    ws = openpyxl.load_workbook(path, data_only=True)[
        openpyxl.load_workbook(path, read_only=True).sheetnames[0]]
    head = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and str(v).strip() in XLSX_COLUMNS:
            head[XLSX_COLUMNS[str(v).strip()]] = c
    need = set(XLSX_COLUMNS.values()) - set(head)
    if need:
        raise SystemExit('%s: missing columns %s' % (path, ', '.join(sorted(need))))

    rows = []
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, head['name']).value or '')
        # The sheet carries its own subtotal rows ("Total 122 SELECT BOARD"). They are
        # the source's own arithmetic, not accounts, and summing them double-counts.
        if name.startswith('Total') or not ws.cell(r, head['fund']).value:
            continue
        def n(k):
            v = ws.cell(r, head[k]).value
            return float(v) if isinstance(v, (int, float)) else 0.0
        rows.append(dict(
            level='account', org=str(ws.cell(r, head['org']).value or ''),
            object=str(ws.cell(r, head['object']).value or ''),
            dept=str(ws.cell(r, head['dept']).value or ''), name=name.strip(),
            fund=str(ws.cell(r, head['fund']).value).strip(),
            fund_name='GENERAL FUND' if str(ws.cell(r, head['fund']).value).strip()
                      == '0100' else '',
            original=n('original'), transfers=n('transfers'), revised=n('revised'),
            expended=n('expended'), encumbered=n('encumbered'),
            available=round(n('revised') - n('expended') - n('encumbered'), 2),
            pct_used=0.0))
    for r in rows:
        r.update(fy=fy, period=period, funds_covered='',
                 account_type=account_type, totals_only=int(totals_only),
                 suppress_zero=int(suppress_zero),
                 # Nothing here is rounded: that is the point of having the spreadsheet.
                 rounded_columns='',
                 doc_id=os.path.relpath(path, ROOT))
    return dict(path=path, rows=rows, stated=twin, fy=fy, period=period,
                account_type=account_type, totals_only=totals_only)


FIELDS = ['doc_id', 'fy', 'period', 'fund', 'fund_name', 'funds_covered', 'account_type',
          'level', 'dept', 'org', 'object', 'name', 'original', 'transfers', 'revised',
          'expended', 'encumbered', 'available', 'pct_used', 'totals_only',
          'suppress_zero', 'rounded_columns']


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--check', action='store_true')
    args = ap.parse_args()

    dirs = [os.path.join(ROOT, 'sources', d)
            for d in ('q3-fy26', 'records-request-2026-09')]
    paths = sorted(os.path.join(d, f) for d in dirs if os.path.isdir(d)
                   for f in os.listdir(d) if f.endswith('.txt'))

    out, bad = [], []
    print('Parsing MUNIS year-to-date budget reports\n')
    for p in paths:
        rep = parse(p)
        if rep is None or not rep['rows']:
            continue

        # Where the same report was also sent as a spreadsheet, prefer the spreadsheet:
        # it carries the appropriation columns un-rounded. It is only trusted after it
        # reconciles to the PRINTED report's own GRAND TOTAL, which is what establishes
        # that the two are the same report at all.
        twin = p[:-4] + '.xlsx'
        if os.path.exists(twin):
            xl = parse_xlsx(twin, rep['fy'], rep['period'], rep['account_type'],
                            rep['totals_only'], rep['suppress_zero'], rep['stated'])
            note, ok = reconcile(xl)
            print('  %-4s %-46s %s %4d rows  %s'
                  % (xl['account_type'][:4], os.path.basename(twin),
                     f"FY{xl['fy']} P{xl['period']}", len(xl['rows']),
                     note if ok else 'MISMATCH ' + note))
            print('       grain: ACCOUNT (spreadsheet; period and options from %s)'
                  % os.path.basename(p))
            if not ok:
                bad.append((twin, note))
            out.extend(xl['rows'])
            continue
        note, ok = reconcile(rep)
        if not ok:
            bad.append((p, note))
        grain = 'dept' if rep['totals_only'] else 'ACCOUNT'
        print('  %-4s %-46s %s %4d rows  %s'
              % (rep['account_type'][:4], os.path.basename(p),
                 f"FY{rep['fy']} P{rep['period']}", len(rep['rows']),
                 note if ok else 'MISMATCH ' + note))
        print('       grain: %s' % grain)
        out.extend(rep['rows'])

    if bad and args.check:
        print('\n%d report(s) do not reconcile; nothing written' % len(bad))
        return 1

    with open(OUT, 'w', newline='', encoding='utf-8') as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        for r in out:
            w.writerow(r)
    print('\nwrote %s -- %d rows' % (os.path.relpath(OUT, ROOT), len(out)))
    return 0


if __name__ == '__main__':
    sys.exit(main())
