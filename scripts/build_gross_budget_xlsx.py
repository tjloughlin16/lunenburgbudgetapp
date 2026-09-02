"""Build the gross school budget as a spreadsheet, in the district's own shape.

    python3 scripts/build_gross_budget_xlsx.py
    python3 scripts/build_gross_budget_xlsx.py --fy 2026

Writes `fy28/public/data/gross-school-budget-fy<YY>.xlsx`.

WHY A SPREADSHEET AND NOT ANOTHER PAGE

Because this is how the budget is actually read. The district publishes a workbook, the
Finance Committee works from a workbook, and a resident who wants to check a line opens a
workbook. A web page is a different artefact for a different reader; this is the one that
sits beside theirs.

WHAT IT IS

The district's budget with two things added that its own version cannot show:

  1.  what was actually SPENT, from the town's ledger
  2.  what OTHER money paid for the same thing -- grants, revolving funds, fees

and, where either is not held, a cell that says so in those words rather than being left
blank. **A blank cell reads as zero.** That is the single most important thing about this
file: the emptiness has to be as legible as the figures, because the whole reason the
district's own budget misleads is that it shows a net number with nothing marking it net.

STRUCTURE, DELIBERATELY THEIRS

Same sections, same 78 function groups, same line names, same order. A reader should be
able to put this beside the district's document and follow it row for row. Where we add a
column it goes to the right of theirs, never in among them.

FIVE SHEETS

  Gross budget      their structure, their columns, plus ours
  Ledger FY26       the 258 account rows we actually hold, which is a DIFFERENT structure
  Funds outside     money spent on the schools from other funds, attributable to nothing
  What is missing   every gap, what would fill it, and who to ask
  Reconciliation    proof that the net column still ties to the published appropriation

THE TWO STRUCTURES DO NOT JOIN, AND THE FILE SAYS SO

The district names its lines; the ledger codes its accounts; no published document maps
one to the other. So the ledger is a separate sheet rather than a column, because putting
it in a column would imply a row-for-row correspondence that nobody has established.
Matching `CLASS ADS` to "Classified Ads" is a judgement, and a plausible name match
quoted as a fact is the error this project exists to avoid.
"""
import argparse
import csv
import os
import sqlite3
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB = os.path.join(ROOT, 'sources', 'data', 'lunenburg.db')
LINES = os.path.join(ROOT, 'sources', 'data', 'lps-budget-lines.csv')
OUT_DIR = os.path.join(ROOT, 'fy28', 'public', 'data')

INK = '1B1B1B'
MUTED = '6B6B6B'
RULE = 'D9D3C7'
BAND = 'F4F1EA'
GAP_FILL = PatternFill('solid', fgColor='FDF0D5')      # amber: known missing
GAP_INK = '8A5A00'
HEAD_FILL = PatternFill('solid', fgColor='1B1B1B')
GRP_FILL = PatternFill('solid', fgColor='EDE7DA')
MONEY = '#,##0;[Red]-#,##0'
MONEY2 = '#,##0.00;[Red]-#,##0.00'

thin = Side(style='thin', color=RULE)
BOX = Border(bottom=thin)

# The district's own columns, in the district's own order, under the district's own
# headings. Nothing here is renamed: a reader has to be able to follow this beside their
# document, and a helpful rename is a silent edit to somebody else's budget.
THEIRS = [
    ('fy23_actual', 'FY23', 'ACTUALS'),
    ('fy24_actual', 'FY24', 'ACTUALS'),
    ('fy25_actual', 'FY25', 'ACTUALS'),
    ('fy25_budget', 'FY25', 'Budget'),
    ('fy26_final', 'FY26', 'FINAL BUDGET'),
    ('fy26_actual_td', 'FY26', 'Actuals to date'),
    ('fy26_encumb_td', 'FY26', 'Encumbrances to date'),
    ('fy27_balanced', 'FY27', 'Balanced Proposed'),
]

# What we add. Each carries the state of the thing it reports, because a column of blanks
# and a column of zeroes look identical and mean opposite things.
OURS = [
    ('town_spent', 'FY26', 'Town actually spent', 'ledger'),
    ('grants', 'FY26', 'Paid by grants', 'not held'),
    ('revolving', 'FY26', 'Paid by revolving funds & fees', 'not held'),
    ('other', 'FY26', 'Paid by other sources', 'not held'),
    ('gross', 'FY26', 'GROSS — what it cost', 'not computable'),
    ('town_share', 'FY26', 'Town’s share of gross', 'not computable'),
]


def money(ws, cell, v, fmt=MONEY):
    cell.value = v
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal='right')


def gap(cell, label='not held'):
    """A gap, written down. Never a blank -- a blank reads as zero."""
    cell.value = label
    cell.fill = GAP_FILL
    cell.font = Font(size=8, italic=True, color=GAP_INK)
    cell.alignment = Alignment(horizontal='center')


def num(v):
    if v in (None, '', '-'):
        return None
    try:
        return float(v)
    except ValueError:
        return None


def sheet_gross(wb, rows, fy):
    ws = wb.create_sheet('Gross budget')
    ws.sheet_view.showGridLines = False

    ws['A1'] = 'LUNENBURG PUBLIC SCHOOLS — GROSS BUDGET'
    ws['A1'].font = Font(bold=True, size=15, color=INK)
    ws['A2'] = ('The district’s own budget, with what was actually spent and what other '
                'money paid for it. Built %s by scripts/build_gross_budget_xlsx.py.'
                % date.today().isoformat())
    ws['A2'].font = Font(size=9, color=MUTED)
    ws['A3'] = ('A BUDGET LINE IS NET. It is what the town must raise after grants, fees '
                'and reimbursements have paid for part of the thing — so a $20,000 line '
                'can be a $220,000 line. Amber cells are money we know exists and cannot '
                'yet attribute. THEY ARE NOT ZERO.')
    ws['A3'].font = Font(size=9, bold=True, color=GAP_INK)
    ws['A4'] = ('Columns A–L are the district’s, unchanged and in their order. Columns M '
                'onward are ours. FY26 figures are period 12 — June, books not closed.')
    ws['A4'].font = Font(size=9, color=MUTED)

    r = 6
    hdr1, hdr2 = r, r + 1
    ws.cell(hdr2, 1, 'SECTION').font = Font(bold=True, size=8, color='FFFFFF')
    ws.cell(hdr2, 2, 'FUNCTION GROUP').font = Font(bold=True, size=8, color='FFFFFF')
    ws.cell(hdr2, 3, 'LINE ITEM').font = Font(bold=True, size=8, color='FFFFFF')
    ws.cell(hdr2, 4, 'THEIR ROW').font = Font(bold=True, size=8, color='FFFFFF')
    col = 5
    for _, y, k in THEIRS:
        ws.cell(hdr1, col, y).font = Font(bold=True, size=8, color='FFFFFF')
        ws.cell(hdr2, col, k).font = Font(bold=True, size=8, color='FFFFFF')
        col += 1
    ws.cell(hdr1, col, 'ADDED BY THIS PROJECT →').font = Font(
        bold=True, size=8, color='FFD79B')
    for _, y, k, _st in OURS:
        ws.cell(hdr2, col, k).font = Font(bold=True, size=8, color='FFD79B')
        col += 1
    last_col = col - 1
    for c in range(1, last_col + 1):
        for rr in (hdr1, hdr2):
            ws.cell(rr, c).fill = HEAD_FILL
            ws.cell(rr, c).alignment = Alignment(
                horizontal='center' if c > 4 else 'left', wrap_text=True, vertical='bottom')

    r = hdr2 + 1
    group = None
    for row in rows:
        if row['kind'] != 'line' or not row['line_item']:
            continue
        g = row['function_group'] or ''
        if g and g != group:                      # the district's own group header
            group = g
            ws.cell(r, 2, g).font = Font(bold=True, size=9, color=INK)
            for c in range(1, last_col + 1):
                ws.cell(r, c).fill = GRP_FILL
            r += 1

        ws.cell(r, 1, row['section']).font = Font(size=8, color=MUTED)
        ws.cell(r, 3, row['line_item']).font = Font(size=9)
        ws.cell(r, 4, int(row['row'])).font = Font(size=8, color=MUTED)
        c = 5
        for key, _, _ in THEIRS:
            v = num(row[key])
            if v is None:
                ws.cell(r, c).value = '—'
                ws.cell(r, c).font = Font(size=8, color=MUTED)
                ws.cell(r, c).alignment = Alignment(horizontal='center')
            else:
                money(ws, ws.cell(r, c), v)
                ws.cell(r, c).font = Font(size=9)
            c += 1
        # Ours. Every one is a gap today, and each says which gap it is.
        gap(ws.cell(r, c), 'no crosswalk')
        gap(ws.cell(r, c + 1))
        gap(ws.cell(r, c + 2))
        gap(ws.cell(r, c + 3))
        gap(ws.cell(r, c + 4), 'not computable')
        gap(ws.cell(r, c + 5), 'not computable')
        for cc in range(1, last_col + 1):
            ws.cell(r, cc).border = BOX
        r += 1

    ws.freeze_panes = ws.cell(hdr2 + 1, 5)
    for c, wdt in ((1, 11), (2, 30), (3, 40), (4, 9)):
        ws.column_dimensions[get_column_letter(c)].width = wdt
    for c in range(5, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.row_dimensions[hdr2].height = 30
    return ws, r


def sheet_ledger(wb, db, fy, period=12):
    ws = wb.create_sheet('Ledger FY%d' % fy)
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'THE TOWN’S LEDGER — SCHOOL DEPARTMENT, ACCOUNT LEVEL'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = ('Period %d. This is a DIFFERENT structure from the sheet before it. '
                'The district names its lines; the ledger codes its accounts; no '
                'published document maps one to the other. They are kept apart on '
                'purpose — putting them side by side would imply a row-for-row match '
                'nobody has established.' % period)
    ws['A2'].font = Font(size=9, color=GAP_INK, bold=True)

    head = ['ORG', 'OBJECT', 'ACCOUNT', 'APPROPRIATED', 'TRANSFERS', 'REVISED',
            'SPENT', 'ENCUMBERED', 'UNSPENT', '% USED']
    for i, h in enumerate(head, 1):
        c = ws.cell(4, i, h)
        c.font = Font(bold=True, size=8, color='FFFFFF')
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal='center' if i > 3 else 'left')
    r = 5
    for a in db.execute("""SELECT a.org, a.object, a.name, l.original, l.transfers,
                                  l.revised, l.expended, l.encumbered, l.available
                           FROM ledger_snapshot l JOIN account a USING (account_id)
                           WHERE l.fy=? AND l.period=? AND a.dept='300'
                             AND a.level='account'
                           ORDER BY a.org, a.object""", (fy, period)):
        ws.cell(r, 1, a[0]).font = Font(size=8, name='Menlo')
        ws.cell(r, 2, a[1]).font = Font(size=8, name='Menlo')
        ws.cell(r, 3, a[2]).font = Font(size=9)
        for i, v in enumerate(a[3:], 4):
            money(ws, ws.cell(r, i), float(v or 0), MONEY2)
            ws.cell(r, i).font = Font(size=9)
        rev, sp = float(a[5] or 0), float(a[6] or 0)
        pc = ws.cell(r, 10)
        pc.value = (sp / rev) if rev else None
        pc.number_format = '0%'
        pc.alignment = Alignment(horizontal='right')
        if rev and sp > rev:
            pc.font = Font(size=9, bold=True, color='B3261E')
        elif rev and sp == 0:
            pc.font = Font(size=9, bold=True, color=GAP_INK)
        else:
            pc.font = Font(size=9)
        if r % 2 == 0:
            for c in range(1, 11):
                ws.cell(r, c).fill = PatternFill('solid', fgColor=BAND)
        r += 1
    ws.freeze_panes = 'D5'
    for c, wdt in ((1, 11), (2, 10), (3, 16)):
        ws.column_dimensions[get_column_letter(c)].width = wdt
    for c in range(4, 11):
        ws.column_dimensions[get_column_letter(c)].width = 15
    return ws


def sheet_funds(wb, db, fy):
    ws = wb.create_sheet('Funds outside')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'MONEY SPENT ON THE SCHOOLS FROM OUTSIDE THE GENERAL FUND'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = ('Every dollar here paid for real staff or real programmes and belongs '
                'against one of the lines in the first sheet. WE CANNOT SAY WHICH. '
                'Spreading it across lines in proportion to their size would look right '
                'and be invented.')
    ws['A2'].font = Font(size=9, bold=True, color=GAP_INK)

    for i, h in enumerate(['FUND', 'NAME', 'RESTRICTED TO', 'REVENUE IN', 'SPENT',
                           'BALANCE', 'WHICH LINES'], 1):
        c = ws.cell(4, i, h)
        c.font = Font(bold=True, size=8, color='FFFFFF')
        c.fill = HEAD_FILL
    r = 5
    total = 0.0
    for f in db.execute("""SELECT fa.fund, f.name, f.restriction, fa.revenue,
                                  fa.salaries + fa.expenditure AS spent,
                                  fa.closing_balance
                           FROM fund_activity fa LEFT JOIN fund f ON f.fund = fa.fund
                           WHERE fa.fy=? AND (fa.salaries + fa.expenditure) > 0
                           ORDER BY spent DESC""", (fy,)):
        ws.cell(r, 1, f[0]).font = Font(size=9, name='Menlo')
        ws.cell(r, 2, f[1] or '').font = Font(size=9)
        if f[2]:
            ws.cell(r, 3, f[2]).font = Font(size=8, color=MUTED)
        else:
            gap(ws.cell(r, 3), 'not stated in any document we hold')
        for i, v in enumerate(f[3:], 4):
            money(ws, ws.cell(r, i), float(v or 0), MONEY2)
            ws.cell(r, i).font = Font(size=9)
        gap(ws.cell(r, 7))
        total += float(f[4] or 0)
        r += 1
    ws.cell(r, 2, 'TOTAL, attributable to no line').font = Font(bold=True, size=10)
    money(ws, ws.cell(r, 5), total, MONEY2)
    ws.cell(r, 5).font = Font(bold=True, size=10)
    for c, wdt in ((1, 9), (2, 34), (3, 34), (4, 15), (5, 15), (6, 15), (7, 14)):
        ws.column_dimensions[get_column_letter(c)].width = wdt
    return ws, total


def sheet_missing(wb, fy):
    ws = wb.create_sheet('What is missing')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'WHAT IS MISSING, AND WHAT WOULD FILL IT'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = ('Read this before concluding anything from a blank. Every amber cell in '
                'this workbook is one of the rows below.')
    ws['A2'].font = Font(size=9, color=MUTED)
    for i, h in enumerate(['COLUMN / CELL', 'WHAT IT WOULD SHOW', 'WHY IT IS EMPTY',
                           'WHAT WOULD FILL IT', 'WHO HAS IT'], 1):
        c = ws.cell(4, i, h)
        c.font = Font(bold=True, size=8, color='FFFFFF')
        c.fill = HEAD_FILL
    gaps = [
        ('Town actually spent', 'What the town paid against this line',
         'The district names its lines; the ledger codes its accounts; no published '
         'document maps one to the other, so no ledger figure can be attached to a named '
         'line. The account-level figures are on the Ledger sheet.',
         'A crosswalk from budget line to ledger account, established one row at a time '
         'with the evidence for each. The account descriptions in the FY26 ledger are the '
         'first material for building it.', 'Nobody publishes it'),
        ('Paid by grants', 'Grant money that paid for part of this line',
         'Grant funds are outside Fund 0100 and every expenditure report held is Fund '
         '0100 only.',
         'The MUNIS year-to-date budget report (glytdbud) run for the school grant funds '
         '— 2813 and 2814 (IDEA #240), 2705/2713 (Title I), 2728 (Title IIA), 2750 '
         '(Title IV), 2778 (SOA) — with Print totals only: N.', 'Town Accountant'),
        ('Paid by revolving funds & fees', 'Fee and revolving fund money on this line',
         'Same reason: outside Fund 0100.',
         'The same report for 1301 (athletics), 1305 (after school), 1306 (facilities '
         'use), 1308 (school choice), 1311 (gifts), 1312 (extended day), 2200 (lunch).',
         'Town Accountant'),
        ('Paid by other sources', 'Circuit breaker, gifts, anything else',
         'Not separated in anything held.',
         'DESE End of Year Financial Report, Schedule 1, which separates revenues and '
         'expenditures BY SOURCE OF FUNDS. DESE does not publish it per district; the '
         'district files it and holds a copy.', 'Lunenburg Public Schools'),
        ('GROSS — what it cost', 'The real cost of the thing, all sources',
         'It is the sum of the four columns to its left and three of them are empty.',
         'Any of the above. Each one fills part of it.', '—'),
        ('Town’s share of gross', 'What share of the cost falls on the levy',
         'Cannot be computed without the gross.', 'Any of the above.', '—'),
        ('FY26 figures generally', 'A closed year',
         'The FY26 report held is period 12 — June, with the books open. Period 13 is '
         'the year-end close after purchase orders are cleared, which moved the FY25 '
         'school figure by $21,770.53.',
         'The same report at Year/Period 2026/13, plus the purchase orders closed against '
         'the year after its initial close.', 'Town Accountant'),
    ]
    r = 5
    for g in gaps:
        for i, v in enumerate(g, 1):
            c = ws.cell(r, i, v)
            c.font = Font(size=9, bold=(i == 1))
            c.alignment = Alignment(wrap_text=True, vertical='top')
        if r % 2 == 0:
            for i in range(1, 6):
                ws.cell(r, i).fill = PatternFill('solid', fgColor=BAND)
        ws.row_dimensions[r].height = 62
        r += 1
    for c, wdt in ((1, 26), (2, 34), (3, 46), (4, 52), (5, 22)):
        ws.column_dimensions[get_column_letter(c)].width = wdt
    return ws


def sheet_recon(wb, db, rows, fy, outside, period=12):
    ws = wb.create_sheet('Reconciliation')
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'DOES THIS STILL TIE TO THE PUBLISHED BUDGET?'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = ('If it does not, this workbook is a parallel invention rather than a '
                'check on the district’s. Every figure below is computed, not typed.')
    ws['A2'].font = Font(size=9, color=MUTED)

    wb_total = sum(num(r['fy26_final']) or 0 for r in rows if r['kind'] == 'line')
    # ALWAYS filter on period. A fiscal year is held at more than one period -- FY26 at
    # period 9 as a department rollup and at period 12 at account level -- so a sum
    # without it counts the same appropriation twice. That is exactly what this
    # reconciliation caught on its first run: dept 301's $40,000, doubled.
    led = db.execute("""SELECT SUM(original), SUM(expended), SUM(encumbered),
                               SUM(available)
                        FROM ledger_snapshot l JOIN account a USING (account_id)
                        WHERE l.fy=? AND l.period=? AND a.dept='300'
                          AND a.level='account'""", (fy, period)).fetchone()
    dept301 = db.execute("""SELECT SUM(original) FROM ledger_snapshot l
                            JOIN account a USING (account_id)
                            WHERE l.fy=? AND l.period=? AND a.dept='301'""",
                         (fy, period)).fetchone()[0] or 0

    checks = [
        ('The district’s FY26 budget, summed from its own line rows', wb_total, ''),
        ('The town’s FY26 appropriation, school department (300)', float(led[0]), ''),
        ('  plus school non-recurring (301)', float(dept301), ''),
        ('  = the town’s appropriation to the schools', float(led[0]) + dept301,
         'ties to the district’s own total to within a dollar' if
         abs(wb_total - (float(led[0]) + dept301)) < 2 else
         'DOES NOT TIE — investigate before publishing anything from this file'),
        ('', None, ''),
        ('Spent from the general fund, period 12', float(led[1]), ''),
        ('Still encumbered', float(led[2]), ''),
        ('Unspent', float(led[3]), 'NOT a surplus — period 12, books open'),
        ('', None, ''),
        ('Known spent outside the general fund', outside,
         'attributable to no line — see Funds outside'),
        ('GROSS FLOOR (town spending + known outside)', float(led[1]) + outside,
         'A FLOOR, never a total. Grant funds are only partly visible.'),
    ]
    r = 4
    for label, val, note in checks:
        if val is None:
            r += 1
            continue
        ws.cell(r, 1, label).font = Font(size=10, bold=label.startswith(('The district',
                                                                        'GROSS')))
        money(ws, ws.cell(r, 2), val, MONEY2)
        ws.cell(r, 2).font = Font(size=10, bold=True)
        if note:
            c = ws.cell(r, 3, note)
            c.font = Font(size=9, italic=True,
                          color='B3261E' if 'NOT TIE' in note else GAP_INK)
        r += 1
    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 60
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--fy', type=int, default=2026)
    a = ap.parse_args()

    rows = list(csv.DictReader(open(LINES, encoding='utf-8')))
    db = sqlite3.connect(DB)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_gross(wb, rows, a.fy)
    sheet_ledger(wb, db, a.fy)
    _, outside = sheet_funds(wb, db, a.fy)
    sheet_missing(wb, a.fy)
    sheet_recon(wb, db, rows, a.fy, outside)

    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, 'gross-school-budget-fy%d.xlsx' % a.fy)
    wb.save(out)
    print('wrote %s  (%.0f KB)' % (os.path.relpath(out, ROOT),
                                   os.path.getsize(out) / 1024))
    print('  sheets: %s' % ', '.join(wb.sheetnames))
    return 0


if __name__ == '__main__':
    sys.exit(main())
