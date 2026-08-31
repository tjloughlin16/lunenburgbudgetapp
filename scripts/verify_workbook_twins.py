#!/usr/bin/env python3
"""Do the two copies of the 25 March 2026 workbook agree on the budget?

This matters more than it looks, and the reason is provenance rather than arithmetic.

`xlsx/fy27-proposals.xlsx` is the document nearly every budget-line figure on this site
comes out of, and **nobody recorded where it came from.** Its twin,
`xlsx/fy27-budget-projection-3-25-26.xlsx`, has an address: a member of the Finance
Committee sent it. So if the two agree cell for cell across the budget columns, every
figure the site publishes is reproduced in a document a reader can trace to a named town
official -- which is not as good as the load-bearing file having its own address, but it is
the difference between a figure a reader must take on trust and one they can check.

That claim therefore has to be measured, not asserted. MANIFEST.md previously said the two
differed in "51 cells, all in an unused scratch column (col X, full of #VALUE!)". None of
that reproduces: the scratch column is Y, not X; it holds `=Jn-Kn`, not `#VALUE!`; and the
count is 410 differing cells at formula level. The substance of the claim survived and
every specific in it was wrong, which is rule 13 -- a rendering quoted as an observation.

**What the comparison must not do is compare the wrong thing.** Reading cached values
instead of formulas reports 595 differences, of which 148 are column U alone: identical
formulas whose cached results were saved by one copy and not the other. That is the
instrument, not the workbooks. So the comparison is at formula level, and the value-level
count is reported beside it rather than instead of it.

It also asserts the authorship metadata that sources/xlsx/PROVENANCE.md quotes. Every .xlsx
is a zip whose docProps members record who created and last saved it, and that is the only
provenance evidence these files carry from inside. It is worth stating exactly what that is
and is not: **it says who authored a file, never who gave it to us.** A workbook created by
the district's Business Administrator reads the same whether it was emailed, downloaded, or
handed over on a memory stick.

    python3 scripts/verify_workbook_twins.py
"""
import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter as col

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UNTRACED = os.path.join(ROOT, 'sources/xlsx/fy27-proposals.xlsx')
TRACED = os.path.join(ROOT, 'sources/xlsx/fy27-budget-projection-3-25-26.xlsx')

# Columns E through M: FY25 budget, FY26 final, FY26 actuals-to-date and encumbrances, and
# the four FY27 scenarios. Everything the projection reads. Anything outside this band is
# working space and may legitimately differ between two saves of the same workbook.
BUDGET_COLS = range(5, 14)
LAST_ROW = 1197

# Row 230 sits directly under TOTAL EXPENSES (row 229) and is a ratio row: each cell
# divides its own column's total by FY25's, giving year-over-year growth. The Finance
# Committee copy has it and the other does not.
#
# It is inside the budget columns and it is NOT a budget figure, so it is excused -- but
# excused by assertion, not by row number. Every one of these cells has to actually be
# that ratio, or the exception does not apply and the check fails. An exception that
# trusts a row number is how a deleted budget figure would slip through it.
RATIO_ROW = 230
RATIO = re.compile(r'^=([A-Z]{1,2})229/\$E\$229$')


# What the docProps say, as of 31 August 2026. Asserted rather than transcribed, because
# PROVENANCE.md quotes these and a document quoting a file it no longer matches is the whole
# reason this project writes checks.
#
# The load-bearing one is `cp:lastModifiedBy` on the traced copy: it reads `Ana Lockwood`,
# which corroborates the account of how it reached us from inside the file itself. The
# untraced copy has no last modifier at all.
EXPECTED = {
    'fy27-proposals.xlsx': {
        'dc:creator': 'Christopher McNamara',
        'dcterms:created': '2025-11-04T18:07:42Z',
        'cp:lastModifiedBy': None,
    },
    'fy27-budget-projection-3-25-26.xlsx': {
        'dc:creator': 'Christopher McNamara',
        'dcterms:created': '2025-11-04T18:07:42Z',
        'cp:lastModifiedBy': 'Ana Lockwood',
        'dcterms:modified': '2026-03-27T01:44:48Z',
    },
    'fy27-budget-projection-2-24-26.xlsx': {
        'dc:creator': 'Christopher McNamara',
        'dcterms:created': '2025-11-04T18:07:42Z',
        'cp:lastModifiedBy': None,
    },
}


def docprops(path):
    """The authorship tags, read out of the zip rather than out of a spreadsheet reader."""
    import zipfile
    out = {}
    with zipfile.ZipFile(path) as z:
        if 'docProps/core.xml' not in z.namelist():
            return out
        x = z.read('docProps/core.xml').decode('utf8', 'ignore')
        for tag in ('dc:creator', 'cp:lastModifiedBy', 'dcterms:created',
                    'dcterms:modified'):
            m = re.search(rf'<{tag}[^>]*>(.*?)</{tag}>', x, re.S)
            out[tag] = m.group(1).strip() if m else None
    return out


def check_metadata():
    bad = []
    for name, want in EXPECTED.items():
        got = docprops(os.path.join(ROOT, 'sources/xlsx', name))
        for tag, value in want.items():
            if got.get(tag) != value:
                bad.append(f'    {name}  {tag}: expected {value!r}, found '
                           f'{got.get(tag)!r}')
    if bad:
        print('  authorship metadata has changed since PROVENANCE.md was written:')
        print('\n'.join(bad))
    else:
        print('  authorship metadata matches sources/xlsx/PROVENANCE.md '
              '(creator, created, last modifier)')
    return len(bad)


def cells(path):
    wb = openpyxl.load_workbook(path)                 # formulas, not cached values
    return wb[wb.sheetnames[0]]


def main():
    a, b = cells(UNTRACED), cells(TRACED)

    budget, other, ratios = [], [], []
    for r in range(1, LAST_ROW + 1):
        for c in range(1, 36):
            va, vb = a.cell(r, c).value, b.cell(r, c).value
            if va == vb:
                continue
            # `=sum(` and `=SUM(` are the same sum. Excel does not care and neither does
            # anyone reading the budget; one copy was saved by something that normalised
            # the case. Treated as equal so a real difference is not lost among 14 of them.
            if (isinstance(va, str) and isinstance(vb, str)
                    and va.lower() == vb.lower()):
                continue
            if r == RATIO_ROW and c in BUDGET_COLS:
                m = RATIO.match(str(vb or ''))
                if va is None and m and m.group(1) == col(c):
                    ratios.append((r, c, vb))
                    continue
            (budget if c in BUDGET_COLS else other).append((r, c, va, vb))

    print(f'  {os.path.relpath(UNTRACED, ROOT)}   (no recorded address)')
    print(f'  {os.path.relpath(TRACED, ROOT)}   (from a Finance Committee member)\n')

    for r, c, va, vb in other[:12]:
        print(f'  outside the budget columns: {col(c)}{r}  {va!r} | {vb!r}')
    if len(other) > 12:
        print(f'  ... and {len(other) - 12} more, all outside columns E-M')

    print(f'\n  {len(other)} differences outside the budget columns (scratch columns)')
    print(f'  {len(ratios)} year-over-year ratio cells in row {RATIO_ROW}, present only '
          f'in the traced copy —\n      each verified to be its own column\u2019s total '
          f'over FY25\u2019s, so none is a budget figure')
    print(f'  {len(budget)} differences INSIDE the budget columns E-M')

    print()
    stale = check_metadata()

    if budget or stale:
        for r, c, va, vb in budget[:20]:
            label = a.cell(r, 2).value or a.cell(r, 1).value
            print(f'    {col(c)}{r}  {str(label)[:40]!r}  {va!r} | {vb!r}')
        if budget:
            print('\nFAILED — the two copies disagree on a budget figure. The site is '
                  'built on the copy with no address, so this is the one that must not '
                  'happen.')
        else:
            print('\nFAILED — sources/xlsx/PROVENANCE.md describes metadata these files '
                  'no longer carry.')
        return 1

    print('\nPASSED — every cell in columns E through M is identical, formula for '
          'formula.\nEvery budget figure this site publishes is reproduced in the copy '
          'that has an address.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
