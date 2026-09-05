"""Is an arriving export safe to ingest? Decided from its headers, without reading a row.

Run this on a file the moment it arrives, BEFORE it is opened in anything, moved into
`sources/`, or shown to any model. It answers one question -- does this file carry a column
that can contain a person's name -- and it answers it from the header row alone.

WHY HEADERS ONLY

Because the check must not become the disclosure. A scan that reads every cell looking for
names has read every name, and if that scan runs anywhere but this machine, the exposure has
already happened. A header row is metadata: it says a vendor column EXISTS without revealing
one value in it. That is enough to decide, and it is the least this can look at.

So this script never prints a cell value, and never returns one. It reads row 1 and stops.

WHAT IT KNOWS

The MUNIS "Journal Detail Export" that the Town has sent before carries 25 columns, of which
exactly two can name somebody:

    X  VDR NAME/ITEM DESC   -- on a special education line, often a parent, not a company
    Y  COMMENTS             -- free text, so nothing constrains it

Both are listed in `notes/process/INTAKE-FOR-THE-TOWN.md` as columns to omit at export time. This is
the check that the omission actually happened.

Names are matched loosely on purpose. A column called `VENDOR`, `VDR NAME`, `PAYEE`,
`REMIT TO` or `STUDENT` is the same hazard under a different heading, and an export layout
that changes its labels must fail this check rather than pass it by accident.

    python3 scripts/check_intake_headers.py ~/lunenburg-intake/*.xlsx

Exit 0 = no hazardous column found. Exit 1 = do not ingest; see notes/process/INTAKE.md §10.
"""
import csv
import os
import re
import sys

# Substrings that make a column capable of holding a person. Deliberately broad: a false
# alarm costs an email to the Town, and a miss costs a name in a public archive forever.
HAZARD = [
    'vdr', 'vendor', 'payee', 'remit', 'name', 'student', 'parent', 'guardian',
    'address', 'comment', 'note', 'memo', 'contact', 'phone', 'email', 'ssn', 'dob',
]

# Headers that contain a hazard word but are known to be safe in this export. Each needs a
# reason, because an exception list without reasons becomes a way to silence the check.
ALLOWED = {
    # MUNIS's own account-name column. It is the ACCOUNT's description ('CASH'), not a
    # person's, and every value in the FY24-FY26 exports was an account title.
    'description': 'the account name, not a person -- MUNIS column E',
}


def headers(path):
    """Row one, and nothing else."""
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.csv', '.txt'):
        with open(path, newline='', errors='replace') as fh:
            return next(csv.reader(fh), [])
    if ext in ('.xlsx', '.xlsm'):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        out = []
        for name in wb.sheetnames:
            row = next(wb[name].iter_rows(max_row=1, values_only=True), ())
            out += [f'{name}!{c}' if len(wb.sheetnames) > 1 else str(c)
                    for c in row if c is not None]
        return out
    raise ValueError(f'cannot read headers from {ext} -- see notes/process/INTAKE.md §6')


def main(paths):
    if not paths:
        print(__doc__.strip().splitlines()[-3].strip(), file=sys.stderr)
        return 2
    bad = False
    for path in paths:
        try:
            cols = headers(path)
        except Exception as e:
            print(f'REFUSE  {os.path.basename(path)}: {e}')
            bad = True
            continue
        hits = []
        for col in cols:
            label = re.sub(r'^.*!', '', str(col)).strip()
            key = label.lower()
            if key in ALLOWED:
                continue
            if any(h in key for h in HAZARD):
                hits.append(label)
        name = os.path.basename(path)
        if hits:
            print(f'REFUSE  {name}: {len(cols)} columns, {len(hits)} can name a person')
            for h in hits:
                print(f'          {h}')
            bad = True
        else:
            print(f'  ok    {name}: {len(cols)} columns, none can name a person')
    if bad:
        print('\nDo NOT ingest, open, copy or forward these files. Tell the Town today,')
        print('naming the file and the column, and ask for a re-export without it.')
        print('notes/process/INTAKE.md §10 is the rest of the procedure.')
    return 1 if bad else 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
